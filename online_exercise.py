import secrets
import string
import json
from datetime import date, timedelta

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from openpyxl import load_workbook
from supabase import create_client

try:
    from streamlit_calendar import calendar as st_calendar
except Exception:
    st_calendar = None


st.set_page_config(page_title="LAB33 Training System", layout="wide")


def get_supabase_client():
    return create_client(
        st.secrets["SUPABASE_URL"],
        st.secrets["SUPABASE_ANON_KEY"],
    )


def get_admin_client():
    service_role_key = st.secrets.get("SUPABASE_SERVICE_ROLE_KEY")
    if not service_role_key:
        return None
    return create_client(st.secrets["SUPABASE_URL"], service_role_key)


supabase = get_supabase_client()
admin_supabase = get_admin_client()

APP_URL = st.secrets.get("APP_URL", "http://localhost:8502")

TRAINING_CATEGORIES = [
    "肌力/爆發訓練",
    "速度與敏捷",
    "有氧/無氧訓練",
    "專項訓練",
    "恢復/其他訓練",
    "其他",
]

BLOCK_SECTION_NAMES = [
    "自我筋膜滾動",
    "自我筋膜鬆動",
    "活動度",
    "活化",
    "動態熱身",
    "增強/彈震式訓練",
    "重量訓練",
    "輔助訓練",
    "恢復訓練",
]

MANUAL_BLOCK_SECTION_NAMES = [
    "自我筋膜滾動",
    "活動度",
    "活化",
    "動態熱身",
    "增強/彈震式訓練",
    "重量訓練",
    "輔助訓練",
    "恢復訓練",
]

MANUAL_TEMPLATE_COLUMNS = [
    "exercise_name",
    "sets",
    "reps_or_time",
    "equipment",
    "intensity",
    "weight",
    "rest",
    "video_url",
]

ATHLETE_BLOCK_EXERCISE_COLUMNS = [
    "exercise_name",
    "sets",
    "reps_or_time",
    "equipment",
    "intensity",
    "weight",
    "rest",
    "video_url",
    "notes",
]


def inject_responsive_styles():
    st.markdown(
        """
        <style>
        .block-container {
            max-width: 1180px;
            padding-top: 3rem;
            padding-bottom: 3rem;
        }

        div[data-testid="stDataFrame"] {
            overflow-x: auto;
        }

        div[data-testid="stExpander"] details {
            overflow-wrap: anywhere;
        }

        @media (max-width: 768px) {
            .block-container {
                padding: 1rem 0.85rem 2rem;
            }

            h1 {
                font-size: 2.1rem !important;
                line-height: 1.15 !important;
            }

            h2 {
                font-size: 1.65rem !important;
            }

            h3 {
                font-size: 1.35rem !important;
            }

            div[data-testid="stHorizontalBlock"] {
                flex-wrap: wrap;
                gap: 0.5rem;
            }

            div[data-testid="column"] {
                flex: 1 1 100% !important;
                min-width: 100% !important;
            }

            div[data-testid="stButton"] > button,
            div[data-testid="stFormSubmitButton"] > button {
                width: 100%;
                min-height: 2.75rem;
            }

            div[data-testid="stExpander"] summary {
                line-height: 1.35;
            }

            div[data-testid="stDataFrame"],
            div[data-testid="stTable"] {
                max-width: 100%;
                overflow-x: auto;
            }

            div[data-testid="stDataFrame"] [role="gridcell"],
            div[data-testid="stDataFrame"] [role="columnheader"] {
                white-space: normal !important;
            }
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def write_client():
    return admin_supabase or supabase


def restore_auth_session():
    access_token = st.session_state.get("access_token")
    refresh_token = st.session_state.get("refresh_token")
    if access_token and refresh_token:
        try:
            supabase.auth.set_session(access_token, refresh_token)
        except Exception:
            clear_auth_session()


def clear_auth_session():
    for key in ("auth_user_id", "auth_email", "access_token", "refresh_token"):
        st.session_state.pop(key, None)


def get_query_param(name):
    value = st.query_params.get(name)
    if isinstance(value, list):
        return value[0] if value else None
    return value


def clear_assignment_query_params():
    for key in ("assign_athlete", "assign_date", "assign_ts"):
        if key in st.query_params:
            del st.query_params[key]


def convert_recovery_hash_to_query_params():
    components.html(
        """
        <script>
        const parentUrl = new URL(window.parent.location.href);
        if (parentUrl.hash && parentUrl.hash.includes("access_token=")) {
          const hashParams = parentUrl.hash.substring(1);
          window.parent.location.replace(
            parentUrl.origin + parentUrl.pathname + "?" + hashParams
          );
        }
        </script>
        """,
        height=0,
    )


def password_recovery_page():
    convert_recovery_hash_to_query_params()

    access_token = get_query_param("access_token")
    refresh_token = get_query_param("refresh_token")
    recovery_type = get_query_param("type")

    if not access_token or recovery_type != "recovery":
        return False

    st.title("設定新密碼")
    st.caption("請輸入新的密碼，完成後再用新密碼登入。")

    with st.form("password_recovery_form"):
        new_password = st.text_input("新 Password", type="password")
        confirm_password = st.text_input("確認新 Password", type="password")
        submitted = st.form_submit_button("更新密碼", use_container_width=True)

    if submitted:
        if not new_password:
            st.warning("請輸入新 Password。")
        elif new_password != confirm_password:
            st.warning("兩次輸入的 Password 不一致。")
        else:
            try:
                supabase.auth.set_session(access_token, refresh_token)
                supabase.auth.update_user({"password": new_password})
                clear_auth_session()
                st.query_params.clear()
                st.success("密碼已更新，請回到學員端用新密碼登入。")
            except Exception as exc:
                st.error(f"更新密碼失敗：{exc}")

    return True


def login_page():
    st.title("LAB33 Training System")
    st.caption("學員登入後只能看到自己的課表")

    tab_login, tab_reset = st.tabs(["登入", "忘記密碼"])

    with tab_login:
        with st.form("login_form"):
            email = st.text_input("Email")
            password = st.text_input("Password", type="password")
            submitted = st.form_submit_button("登入", use_container_width=True)

        if submitted:
            try:
                result = supabase.auth.sign_in_with_password(
                    {"email": email, "password": password}
                )
                set_auth_session(result)
                st.rerun()
            except Exception as exc:
                st.error(f"登入失敗：{exc}")

    with tab_reset:
        st.info("如果忘記密碼，請聯絡教練幫你重設臨時密碼。")
        st.write("教練重設後，你可以用臨時密碼登入，系統會要求你立刻設定新密碼。")


def set_auth_session(result):
    st.session_state["auth_user_id"] = result.user.id
    st.session_state["auth_email"] = result.user.email
    st.session_state["access_token"] = result.session.access_token
    st.session_state["refresh_token"] = result.session.refresh_token


def athlete_email_exists(email):
    normalized_email = email.strip().lower()
    data = (
        supabase.table("athletes")
        .select("id")
        .ilike("email", normalized_email)
        .limit(1)
        .execute()
        .data
    )
    return bool(data)


def generate_temp_password(length=12):
    alphabet = string.ascii_letters + string.digits
    body = "".join(secrets.choice(alphabet) for _ in range(length))
    return f"LAB33-{body}"


def render_service_role_help():
    st.error("尚未設定 `SUPABASE_SERVICE_ROLE_KEY`，無法由教練建立或重設學員帳號。")
    st.info("請到 Supabase Project Settings → API 複製 service_role key，加入 `.streamlit/secrets.toml`。")
    st.code(
        """
SUPABASE_SERVICE_ROLE_KEY = "你的-service-role-key"
        """.strip(),
        language="toml",
    )


def render_auth_creation_error(error):
    st.error(f"新增學員失敗：{error}")

    if "database error creating new user" not in str(error).lower():
        return

    st.info("這通常是 Supabase Auth 建立帳號時，被舊的 auth.users trigger 卡住。若你已不使用 public.users，請到 Supabase SQL Editor 檢查並移除舊 trigger。")
    st.code(
        """
-- 1. 先查看 auth.users 上目前有哪些 trigger
select
  trigger_name,
  event_manipulation,
  action_statement
from information_schema.triggers
where event_object_schema = 'auth'
  and event_object_table = 'users';

-- 2. 如果看到會寫入 public.users 的舊 trigger，刪掉它
-- 常見名稱是 on_auth_user_created，但請以第 1 步查到的名稱為準
drop trigger if exists on_auth_user_created on auth.users;

-- 3. 如果有舊 function 也可以一起移除
-- 常見名稱是 public.handle_new_user，但請先確認你已經不需要它
drop function if exists public.handle_new_user();
        """.strip(),
        language="sql",
    )


def create_auth_user_for_athlete(name, email, temp_password):
    if not admin_supabase:
        raise RuntimeError("Missing SUPABASE_SERVICE_ROLE_KEY")

    result = admin_supabase.auth.admin.create_user(
        {
            "email": email,
            "password": temp_password,
            "email_confirm": True,
            "user_metadata": {"name": name},
        }
    )
    return result.user


def reset_athlete_temp_password(user_id, temp_password):
    if not admin_supabase:
        raise RuntimeError("Missing SUPABASE_SERVICE_ROLE_KEY")

    return admin_supabase.auth.admin.update_user_by_id(
        user_id,
        {"password": temp_password},
    )


def find_auth_user_by_email(email):
    if not admin_supabase or not email:
        return None

    normalized_email = email.strip().lower()
    for page in range(1, 11):
        users = admin_supabase.auth.admin.list_users(page=page, per_page=100)
        if not users:
            break
        for user in users:
            if (user.email or "").lower() == normalized_email:
                return user

    return None


def create_or_link_auth_user_for_athlete(name, email, temp_password):
    auth_user = find_auth_user_by_email(email)
    if auth_user:
        reset_athlete_temp_password(auth_user.id, temp_password)
        return auth_user, "Supabase Authentication 裡已有此 Email，已連結並重設臨時密碼。"

    try:
        auth_user = create_auth_user_for_athlete(name, email, temp_password)
        return auth_user, "已建立新的 Supabase Auth 帳號，並產生臨時密碼。"
    except Exception as exc:
        if "already" not in str(exc).lower() and "registered" not in str(exc).lower():
            raise

        auth_user = find_auth_user_by_email(email)
        if not auth_user:
            raise

        reset_athlete_temp_password(auth_user.id, temp_password)
        return auth_user, "Supabase Authentication 裡已有此 Email，已連結並重設臨時密碼。"


def create_or_reset_athlete_temp_password(athlete):
    if not admin_supabase:
        raise RuntimeError("Missing SUPABASE_SERVICE_ROLE_KEY")

    temp_password = generate_temp_password()
    user_id = athlete.get("user_id")

    if has_value(user_id):
        try:
            reset_athlete_temp_password(user_id, temp_password)
            supabase.table("athletes").update(
                {"must_change_password": True}
            ).eq("id", athlete["id"]).execute()
            return temp_password, "已重設臨時密碼。"
        except Exception as exc:
            if "user not found" not in str(exc).lower():
                raise

    auth_user = find_auth_user_by_email(athlete.get("email"))
    if auth_user:
        reset_athlete_temp_password(auth_user.id, temp_password)
        message = "已連結既有 Auth 帳號，並重設臨時密碼。"
    else:
        auth_user = create_auth_user_for_athlete(
            athlete.get("name") or "",
            athlete.get("email"),
            temp_password,
        )
        message = "已建立 Auth 帳號，並產生臨時密碼。"

    supabase.table("athletes").update(
        {
            "user_id": auth_user.id,
            "must_change_password": True,
        }
    ).eq("id", athlete["id"]).execute()

    return temp_password, message


def has_value(value):
    return value is not None and not pd.isna(value) and value != ""


def parse_date_value(value):
    if not has_value(value):
        return date.today()
    try:
        return pd.to_datetime(value).date()
    except Exception:
        return date.today()


def date_to_iso(value):
    if not has_value(value):
        return None
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


def normalize_date_range(value):
    if isinstance(value, (tuple, list)):
        start_date = value[0] if len(value) >= 1 else date.today()
        end_date = value[1] if len(value) >= 2 else start_date
    else:
        start_date = value
        end_date = value

    if not has_value(start_date):
        start_date = date.today()
    if not has_value(end_date):
        end_date = start_date

    return start_date, end_date


def date_range_label(row):
    start_date = row.get("start_date")
    end_date = row.get("end_date")

    if not has_value(start_date):
        start_date = row.get("scheduled_date")
    if not has_value(end_date):
        end_date = start_date

    if not has_value(start_date):
        return ""
    if not has_value(end_date) or str(start_date) == str(end_date):
        return str(start_date)
    return f"{start_date} ~ {end_date}"


def delete_athlete(athlete_id, user_id=None):
    supabase.table("athlete_blocks").delete().eq("athlete_id", athlete_id).execute()
    supabase.table("athletes").delete().eq("id", athlete_id).execute()

    if has_value(user_id) and admin_supabase:
        try:
            admin_supabase.auth.admin.delete_user(str(user_id))
        except Exception as exc:
            message = str(exc).lower()
            if "user not found" in message:
                return "學員資料已刪除；Supabase Auth 帳號原本就不存在，所以已略過 Auth 刪除。"
            raise

    return "已刪除學員。"


def delete_block(block_id):
    block_id = to_plain_int(block_id)
    supabase.table("athlete_blocks").delete().eq("block_id", block_id).execute()
    supabase.table("block_exercises").delete().eq("block_id", block_id).execute()
    supabase.table("block_sections").delete().eq("block_id", block_id).execute()
    supabase.table("blocks").delete().eq("id", block_id).execute()
    return "已刪除板塊，並移除相關的學員課表與詳細內容。"


def render_temp_password_card(password_key, email_key, close_key, extra_keys=None):
    if not st.session_state.get(password_key):
        return

    st.info("請把這組臨時密碼交給學員。此密碼只會顯示在這裡，學員登入後會被要求修改。")
    st.code(
        f"Email: {st.session_state[email_key]}\n"
        f"Temporary Password: {st.session_state[password_key]}"
    )
    if st.button("我已記下，關閉", key=close_key):
        st.session_state.pop(password_key, None)
        st.session_state.pop(email_key, None)
        for key in extra_keys or []:
            st.session_state.pop(key, None)
        st.rerun()


def logout_button():
    if st.sidebar.button("登出", use_container_width=True):
        try:
            supabase.auth.sign_out()
        except Exception:
            pass
        clear_auth_session()
        st.rerun()


def fetch_athletes():
    try:
        data = (
            write_client().table("athletes")
            .select("*")
            .order("id", desc=True)
            .execute()
            .data
        )
        return pd.DataFrame(data or []), None
    except Exception as exc:
        return pd.DataFrame(), exc


def fetch_blocks():
    try:
        data = (
            write_client().table("blocks")
            .select("id,block_code,block_name,goal,training_element,description")
            .order("id", desc=False)
            .execute()
            .data
        )
        return pd.DataFrame(data or []), None
    except Exception as exc:
        return pd.DataFrame(), exc


def fetch_block_sections(block_id):
    try:
        data = (
            write_client().table("block_sections")
            .select("id,block_id,section_name,order_num")
            .eq("block_id", block_id)
            .order("order_num", desc=False)
            .execute()
            .data
        )
        return pd.DataFrame(data or []), None
    except Exception as exc:
        return pd.DataFrame(), exc


def fetch_block_exercises(block_id):
    try:
        data = (
            write_client().table("block_exercises")
            .select("id,block_id,section_id,exercise_name,sets,reps_or_time,equipment,intensity,weight,rest,video_url,order_num,notes")
            .eq("block_id", block_id)
            .order("order_num", desc=False)
            .execute()
            .data
        )
        return pd.DataFrame(data or []), None
    except Exception as exc:
        return pd.DataFrame(), exc


def fetch_athlete_blocks(athlete_id):
    try:
        data = (
            write_client().table("athlete_blocks")
            .select("id,athlete_id,block_id,event_name,cycle_goal,scheduled_date,start_date,end_date,week_num,day_num,training_category,notes,created_at")
            .eq("athlete_id", athlete_id)
            .order("id", desc=True)
            .execute()
            .data
        )
        return pd.DataFrame(data or []), None
    except Exception as exc:
        return pd.DataFrame(), exc


def fetch_athlete_block_exercises(athlete_block_id):
    base_columns = (
        "id,athlete_block_id,section_name,section_order,exercise_name,sets,"
        "reps_or_time,equipment,intensity,weight,rest,video_url,notes,order_num"
    )
    report_columns = f"{base_columns},actual_sets,actual_weight"
    try:
        data = (
            write_client().table("athlete_block_exercises")
            .select(report_columns)
            .eq("athlete_block_id", athlete_block_id)
            .order("section_order", desc=False)
            .order("order_num", desc=False)
            .execute()
            .data
        )
        return pd.DataFrame(data or []), None
    except Exception as exc:
        try:
            data = (
                write_client().table("athlete_block_exercises")
                .select(base_columns)
                .eq("athlete_block_id", athlete_block_id)
                .order("section_order", desc=False)
                .order("order_num", desc=False)
                .execute()
                .data
            )
            df = pd.DataFrame(data or [])
            if not df.empty:
                df["actual_sets"] = ""
                df["actual_weight"] = ""
            return df, None
        except Exception:
            return pd.DataFrame(), exc


def create_athlete(name, email, sport, level, user_id=None, must_change_password=False):
    payload = {
        "name": name,
        "email": email,
        "sport": sport,
        "level": level,
        "must_change_password": must_change_password,
    }
    if user_id:
        payload["user_id"] = user_id

    return (
        supabase.table("athletes")
        .insert(payload)
        .execute()
    )


def cell_text(value):
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.lower() == "nan" else text


def merged_cell_value(sheet, row_index, column_index):
    cell = sheet.cell(row=row_index, column=column_index)
    if cell.value is not None:
        return cell_text(cell.value)

    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return cell_text(
                sheet.cell(
                    row=merged_range.min_row,
                    column=merged_range.min_col,
                ).value
            )

    return ""


def sheet_row_values(sheet, row_index):
    return [
        merged_cell_value(sheet, row_index, column_index)
        for column_index in range(1, sheet.max_column + 1)
    ]


def normalize_section_name(value):
    if not value:
        return ""
    if "自我筋膜" in value:
        return "自我筋膜滾動"
    if "增強" in value or "彈震" in value:
        return "增強/彈震式訓練"
    matched_section = next(
        (section for section in BLOCK_SECTION_NAMES if section == value),
        "",
    )
    return matched_section


def normalize_header_name(value):
    header = cell_text(value).replace(" ", "")
    header_aliases = {
        "動作": "exercise_name",
        "組數": "sets",
        "次數/時間": "reps_or_time",
        "次數／時間": "reps_or_time",
        "工具": "equipment",
        "強度": "intensity",
        "重量": "weight",
        "休息時間": "rest",
        "休息": "rest",
        "影片連結": "video_url",
        "影片": "video_url",
    }
    return header_aliases.get(header, "")


def value_by_header(values, headers, header_key):
    index = headers.get(header_key)
    if index is None or index >= len(values):
        return ""
    return values[index]


def looks_like_block_title(value, sheet_name):
    if not value:
        return False

    excluded_fragments = [
        "LAB33",
        "Sport Performance",
        "週期目標",
        "訓練元素",
        "強調於",
        "核心目標",
        "透過",
        "過程中",
    ]
    if any(fragment in value for fragment in excluded_fragments):
        return False

    if len(value) > 20:
        return False

    return "訓練" in value or value.replace(" ", "") in sheet_name.replace(" ", "")


def extract_block_title(sheet, fallback_name):
    preferred_values = []
    for row_index in range(1, min(3, sheet.max_row) + 1):
        for value in sheet_row_values(sheet, row_index):
            if looks_like_block_title(value, sheet.title):
                preferred_values.append(value)

    if preferred_values:
        return max(preferred_values, key=len)

    top_values = [
        value
        for row_index in range(1, min(5, sheet.max_row) + 1)
        for value in sheet_row_values(sheet, row_index)
        if looks_like_block_title(value, sheet.title)
    ]
    return max(top_values, key=len) if top_values else fallback_name


def to_plain_int(value):
    if value is None or value == "":
        return None
    return int(value)


def clean_manual_cell(value):
    if value is None or pd.isna(value):
        return ""
    return str(value).strip()


def normalize_unique_value(value):
    return clean_manual_cell(value).casefold()


def duplicate_values(values):
    seen = set()
    duplicates = []
    for value in values:
        normalized = normalize_unique_value(value)
        if not normalized:
            continue
        if normalized in seen and value not in duplicates:
            duplicates.append(value)
        seen.add(normalized)
    return duplicates


def validate_new_block_identity(blocks_to_create, existing_blocks_df):
    errors = []
    block_codes = [block.get("block_code") for block in blocks_to_create]
    block_names = [block.get("block_name") for block in blocks_to_create]

    duplicated_codes = duplicate_values(block_codes)
    duplicated_names = duplicate_values(block_names)
    if duplicated_codes:
        errors.append(f"這次匯入/建立中有重複的 Block Code：{', '.join(duplicated_codes)}")
    if duplicated_names:
        errors.append(f"這次匯入/建立中有重複的顯示名稱：{', '.join(duplicated_names)}")

    if existing_blocks_df is not None and not existing_blocks_df.empty:
        existing_codes = {
            normalize_unique_value(value)
            for value in existing_blocks_df.get("block_code", pd.Series(dtype=str)).tolist()
            if normalize_unique_value(value)
        }
        existing_names = {
            normalize_unique_value(value)
            for value in existing_blocks_df.get("block_name", pd.Series(dtype=str)).tolist()
            if normalize_unique_value(value)
        }

        conflicted_codes = [
            code for code in block_codes
            if normalize_unique_value(code) in existing_codes
        ]
        conflicted_names = [
            name for name in block_names
            if normalize_unique_value(name) in existing_names
        ]
        if conflicted_codes:
            errors.append(f"資料庫已存在相同 Block Code：{', '.join(conflicted_codes)}")
        if conflicted_names:
            errors.append(f"資料庫已存在相同顯示名稱：{', '.join(conflicted_names)}")

    return errors


def render_block_identity_errors(errors):
    for error in errors:
        st.error(error)


def split_importable_blocks(parsed_blocks, existing_blocks_df):
    existing_codes = set()
    if existing_blocks_df is not None and not existing_blocks_df.empty:
        existing_codes = {
            normalize_unique_value(value)
            for value in existing_blocks_df.get("block_code", pd.Series(dtype=str)).tolist()
            if normalize_unique_value(value)
        }

    seen_codes = set()
    importable_blocks = []
    skipped_rows = []

    for parsed_block in parsed_blocks:
        block_code = parsed_block["sheet_name"]
        block_name = parsed_block["block_name"]
        normalized_code = normalize_unique_value(block_code)
        reasons = []

        if normalized_code in seen_codes:
            reasons.append("這次檔案中 Block Code 重複")
        if normalized_code in existing_codes:
            reasons.append("資料庫已存在相同 Block Code")

        seen_codes.add(normalized_code)

        if reasons:
            skipped_rows.append(
                {
                    "Block Code": block_code,
                    "顯示名稱": block_name,
                    "跳過原因": "、".join(reasons),
                }
            )
            continue

        importable_blocks.append(parsed_block)

    return importable_blocks, skipped_rows


def empty_manual_template_rows(row_count=4):
    return pd.DataFrame(
        [
            {column: "" for column in MANUAL_TEMPLATE_COLUMNS}
            for _ in range(row_count)
        ]
    )


def manual_template_version():
    if "manual_template_version" not in st.session_state:
        st.session_state["manual_template_version"] = 0
    return st.session_state["manual_template_version"]


def reset_manual_template():
    st.session_state["manual_template_version"] = manual_template_version() + 1


def build_manual_template_block(block_name, goal, training_element, section_tables):
    sections = []

    for section_name in MANUAL_BLOCK_SECTION_NAMES:
        section_df = section_tables.get(section_name)
        if section_df is None:
            continue

        exercises = []
        for _, row in section_df.iterrows():
            exercise_name = clean_manual_cell(row.get("exercise_name"))
            if not exercise_name:
                continue

            exercises.append(
                {
                    "exercise_name": exercise_name,
                    "sets": clean_manual_cell(row.get("sets")),
                    "reps_or_time": clean_manual_cell(row.get("reps_or_time")),
                    "equipment": clean_manual_cell(row.get("equipment")),
                    "intensity": clean_manual_cell(row.get("intensity")),
                    "weight": clean_manual_cell(row.get("weight")),
                    "rest": clean_manual_cell(row.get("rest")),
                    "video_url": clean_manual_cell(row.get("video_url")),
                    "order_num": len(exercises) + 1,
                }
            )

        if exercises:
            sections.append(
                {
                    "section_name": section_name,
                    "exercises": exercises,
                }
            )

    return {
        "sheet_name": block_name,
        "block_name": block_name,
        "goal": goal,
        "training_element": training_element,
        "sections": sections,
    }


def parse_block_sheet(sheet, fallback_name):
    block_name = extract_block_title(sheet, fallback_name)

    goal = ""
    training_element = ""
    for row_index in range(1, min(5, sheet.max_row) + 1):
        values = sheet_row_values(sheet, row_index)
        for index, value in enumerate(values):
            if value == "週期目標":
                goal = next((item for item in values[index + 1:] if item), "")
            if value == "訓練元素":
                training_element = next((item for item in values[index + 1:] if item), "")

    sections = []
    current_section = None
    current_headers = {}
    last_headers = {}
    order_by_section = {}

    for row_index in range(4, sheet.max_row + 1):
        values = sheet_row_values(sheet, row_index)
        non_empty = [value for value in values if value]

        matched_section = next(
            (
                normalize_section_name(value)
                for value in non_empty
                if normalize_section_name(value)
            ),
            None,
        )
        if matched_section:
            current_section = {
                "section_name": matched_section,
                "exercises": [],
            }
            sections.append(current_section)
            current_headers = last_headers.copy()
            order_by_section[matched_section] = 0
            continue

        if "動作" in non_empty:
            current_headers = {}
            for index, value in enumerate(values):
                header_name = normalize_header_name(value)
                if header_name and header_name not in current_headers:
                    current_headers[header_name] = index
            last_headers = current_headers.copy()
            continue

        if not current_section or not current_headers:
            continue

        exercise_index = current_headers.get("exercise_name")
        if exercise_index is None or exercise_index >= len(values):
            continue

        exercise_name = values[exercise_index]
        if not exercise_name:
            continue

        order_by_section[current_section["section_name"]] += 1
        current_section["exercises"].append(
            {
                "exercise_name": exercise_name,
                "sets": value_by_header(values, current_headers, "sets"),
                "reps_or_time": value_by_header(values, current_headers, "reps_or_time"),
                "equipment": value_by_header(values, current_headers, "equipment"),
                "intensity": value_by_header(values, current_headers, "intensity"),
                "weight": value_by_header(values, current_headers, "weight"),
                "rest": value_by_header(values, current_headers, "rest"),
                "video_url": value_by_header(values, current_headers, "video_url"),
                "order_num": order_by_section[current_section["section_name"]],
            }
        )

    return {
        "sheet_name": sheet.title,
        "block_name": block_name,
        "goal": goal,
        "training_element": training_element,
        "sections": sections,
    }


def parse_block_workbook(uploaded_file):
    workbook = load_workbook(uploaded_file, data_only=True)
    parsed_blocks = []

    for sheet in workbook.worksheets:
        parsed_block = parse_block_sheet(sheet, sheet.title)
        exercise_count = sum(len(section["exercises"]) for section in parsed_block["sections"])
        if parsed_block["sections"] or exercise_count:
            parsed_blocks.append(parsed_block)

    return parsed_blocks


def create_block_from_excel(parsed_block, block_code, description):
    block_name = parsed_block.get("block_name") or parsed_block.get("sheet_name") or block_code
    block_response = (
        supabase.table("blocks")
        .insert(
            {
                "block_code": block_code,
                "block_name": block_name,
                "goal": parsed_block.get("goal") or "",
                "training_element": parsed_block.get("training_element") or "",
                "description": description,
            }
        )
        .execute()
    )
    block_id = block_response.data[0]["id"]

    for section_order, section in enumerate(parsed_block["sections"], start=1):
        section_response = (
            supabase.table("block_sections")
            .insert(
                {
                    "block_id": block_id,
                    "section_name": section["section_name"],
                    "order_num": section_order,
                }
            )
            .execute()
        )
        section_id = section_response.data[0]["id"]

        for exercise in section["exercises"]:
            supabase.table("block_exercises").insert(
                {
                    "block_id": block_id,
                    "section_id": section_id,
                    **exercise,
                }
            ).execute()

    return block_id


def assign_block_to_athlete(
    athlete_id,
    block_id,
    event_name,
    cycle_goal,
    start_date,
    end_date,
    week_num,
    day_num,
    training_category,
    notes,
):
    response = (
        write_client().table("athlete_blocks")
        .insert(
            {
                "athlete_id": to_plain_int(athlete_id),
                "block_id": to_plain_int(block_id),
                "event_name": event_name,
                "cycle_goal": cycle_goal,
                "scheduled_date": date_to_iso(start_date),
                "start_date": date_to_iso(start_date),
                "end_date": date_to_iso(end_date),
                "week_num": to_plain_int(week_num),
                "day_num": to_plain_int(day_num),
                "training_category": training_category,
                "notes": notes,
            }
        )
        .execute()
    )
    assignment_id = response.data[0]["id"]
    create_assignment_exercise_snapshot(assignment_id, block_id)
    return response


def update_athlete_block_assignment(
    assignment_id,
    block_id,
    event_name,
    cycle_goal,
    start_date,
    end_date,
    week_num,
    day_num,
    training_category,
    notes,
):
    return (
        write_client().table("athlete_blocks")
        .update(
            {
                "block_id": to_plain_int(block_id),
                "event_name": event_name,
                "cycle_goal": cycle_goal,
                "scheduled_date": date_to_iso(start_date),
                "start_date": date_to_iso(start_date),
                "end_date": date_to_iso(end_date),
                "week_num": to_plain_int(week_num),
                "day_num": to_plain_int(day_num),
                "training_category": training_category,
                "notes": notes,
            }
        )
        .eq("id", to_plain_int(assignment_id))
        .execute()
    )


def delete_athlete_block_assignment(assignment_id):
    write_client().table("athlete_block_exercises").delete().eq(
        "athlete_block_id",
        to_plain_int(assignment_id),
    ).execute()
    return (
        write_client().table("athlete_blocks")
        .delete()
        .eq("id", to_plain_int(assignment_id))
        .execute()
    )


def template_exercise_rows_for_block(block_id):
    sections_df, sections_error = fetch_block_sections(block_id)
    exercises_df, exercises_error = fetch_block_exercises(block_id)
    if sections_error or exercises_error:
        raise sections_error or exercises_error

    rows = []
    if sections_df.empty:
        return rows

    for section_order, section in enumerate(sections_df.to_dict("records"), start=1):
        if exercises_df.empty or "section_id" not in exercises_df.columns:
            continue
        section_exercises = exercises_df[
            exercises_df["section_id"] == section["id"]
        ].copy()
        if section_exercises.empty:
            continue
        section_exercises = section_exercises.sort_values("order_num")

        for order_num, exercise in enumerate(section_exercises.to_dict("records"), start=1):
            rows.append(
                {
                    "section_name": section.get("section_name") or "未命名區段",
                    "section_order": section.get("order_num") or section_order,
                    "exercise_name": exercise.get("exercise_name") or "",
                    "sets": exercise.get("sets") or "",
                    "reps_or_time": exercise.get("reps_or_time") or "",
                    "equipment": exercise.get("equipment") or "",
                    "intensity": exercise.get("intensity") or "",
                    "weight": exercise.get("weight") or "",
                    "rest": exercise.get("rest") or "",
                    "video_url": exercise.get("video_url") or "",
                    "notes": exercise.get("notes") or "",
                    "order_num": exercise.get("order_num") or order_num,
                }
            )
    return rows


def create_assignment_exercise_snapshot(athlete_block_id, block_id):
    rows = template_exercise_rows_for_block(block_id)
    db = write_client()
    db.table("athlete_block_exercises").delete().eq(
        "athlete_block_id",
        to_plain_int(athlete_block_id),
    ).execute()

    for row in rows:
        db.table("athlete_block_exercises").insert(
            {
                "athlete_block_id": to_plain_int(athlete_block_id),
                **row,
            }
        ).execute()


def save_assignment_exercise_tables(athlete_block_id, section_tables):
    db = write_client()
    db.table("athlete_block_exercises").delete().eq(
        "athlete_block_id",
        to_plain_int(athlete_block_id),
    ).execute()

    for section_order, (section_name, section_df) in enumerate(section_tables.items(), start=1):
        if section_df is None:
            continue
        order_num = 0
        for _, row in section_df.iterrows():
            exercise_name = clean_manual_cell(row.get("exercise_name"))
            if not exercise_name:
                continue
            order_num += 1
            db.table("athlete_block_exercises").insert(
                {
                    "athlete_block_id": to_plain_int(athlete_block_id),
                    "section_name": section_name,
                    "section_order": section_order,
                    "exercise_name": exercise_name,
                    "sets": clean_manual_cell(row.get("sets")),
                    "reps_or_time": clean_manual_cell(row.get("reps_or_time")),
                    "equipment": clean_manual_cell(row.get("equipment")),
                    "intensity": clean_manual_cell(row.get("intensity")),
                    "weight": clean_manual_cell(row.get("weight")),
                    "rest": clean_manual_cell(row.get("rest")),
                    "video_url": clean_manual_cell(row.get("video_url")),
                    "notes": clean_manual_cell(row.get("notes")),
                    "order_num": order_num,
                }
            ).execute()


def update_athlete_exercise_report(exercise_id, actual_sets, actual_weight):
    return (
        write_client().table("athlete_block_exercises")
        .update(
            {
                "actual_sets": clean_manual_cell(actual_sets),
                "actual_weight": clean_manual_cell(actual_weight),
            }
        )
        .eq("id", to_plain_int(exercise_id))
        .execute()
    )


def render_block_setup_help(error):
    st.error(f"讀取或新增板塊失敗：{error}")
    st.info("請先到 Supabase SQL Editor 執行下面這段開發用 SQL。")
    st.code(
        """
create table if not exists public.blocks (
  id bigint generated by default as identity primary key,
  block_code text,
  block_name text,
  goal text,
  training_element text,
  description text,
  created_at timestamp with time zone default now()
);

alter table public.blocks add column if not exists block_code text;
alter table public.blocks add column if not exists block_name text;
alter table public.blocks add column if not exists goal text;
alter table public.blocks add column if not exists training_element text;
alter table public.blocks add column if not exists description text;
alter table public.blocks drop column if exists sport;

do $$
begin
  if not exists (
    select 1
    from pg_constraint
    where conrelid = 'public.blocks'::regclass
      and contype = 'p'
  ) then
    alter table public.blocks alter column id set not null;
    alter table public.blocks add constraint blocks_pkey primary key (id);
  end if;
end $$;

create table if not exists public.block_sections (
  id bigint generated by default as identity primary key,
  block_id bigint,
  section_name text,
  order_num integer,
  created_at timestamp with time zone default now()
);

create table if not exists public.block_exercises (
  id bigint generated by default as identity primary key,
  block_id bigint,
  section_id bigint,
  exercise_name text,
  sets text,
  reps_or_time text,
  equipment text,
  intensity text,
  weight text,
  rest text,
  video_url text,
  order_num integer,
  notes text,
  created_at timestamp with time zone default now()
);

alter table public.blocks no force row level security;
alter table public.blocks disable row level security;
alter table public.block_sections no force row level security;
alter table public.block_sections disable row level security;
alter table public.block_exercises no force row level security;
alter table public.block_exercises disable row level security;

drop policy if exists "Development anon manage blocks" on public.blocks;
create policy "Development anon manage blocks"
on public.blocks
for all
to anon
using (true)
with check (true);

grant usage on schema public to anon, authenticated;
grant select, insert, update, delete on table public.blocks to anon, authenticated;
grant select, insert, update, delete on table public.block_sections to anon, authenticated;
grant select, insert, update, delete on table public.block_exercises to anon, authenticated;
grant usage, select on all sequences in schema public to anon, authenticated;

-- 跑完後用這段確認：
select
  relname,
  relrowsecurity,
  relforcerowsecurity
from pg_class
where relname = 'blocks';

select
  policyname,
  roles,
  cmd
from pg_policies
where schemaname = 'public'
  and tablename = 'blocks';
        """.strip(),
        language="sql",
    )


def render_setup_help(error):
    st.error(f"讀取或新增學員失敗：{error}")
    st.info("請先到 Supabase SQL Editor 執行下面這段開發用 SQL。")
    st.code(
        """
create table if not exists public.athletes (
  id bigint generated by default as identity primary key,
  user_id uuid,
  name text,
  email text,
  sport text,
  level text,
  must_change_password boolean default false,
  created_at timestamp with time zone default now()
);

alter table public.athletes add column if not exists user_id uuid;
alter table public.athletes add column if not exists must_change_password boolean default false;

alter table public.athletes disable row level security;

grant usage on schema public to anon, authenticated;
grant select, insert, update, delete on table public.athletes to anon, authenticated;
grant usage, select on all sequences in schema public to anon, authenticated;
        """.strip(),
        language="sql",
    )


def render_block_assignment_setup_help(error):
    st.error(f"讀取或加入板塊失敗：{error}")
    st.info("請先到 Supabase SQL Editor 執行下面這段開發用 SQL。")
    st.code(
        """
create table if not exists public.athlete_blocks (
  id bigint generated by default as identity primary key,
  athlete_id bigint references public.athletes(id) on delete cascade,
  block_id bigint,
  event_name text,
  cycle_goal text,
  scheduled_date date,
  start_date date,
  end_date date,
  week_num integer,
  day_num integer,
  training_category text,
  notes text,
  created_at timestamp with time zone default now()
);

alter table public.athlete_blocks add column if not exists event_name text;
alter table public.athlete_blocks add column if not exists cycle_goal text;
alter table public.athlete_blocks add column if not exists scheduled_date date;
alter table public.athlete_blocks add column if not exists start_date date;
alter table public.athlete_blocks add column if not exists end_date date;
alter table public.athlete_blocks add column if not exists training_category text;

create table if not exists public.athlete_block_exercises (
  id bigint generated by default as identity primary key,
  athlete_block_id bigint references public.athlete_blocks(id) on delete cascade,
  section_name text,
  section_order integer,
  exercise_name text,
  sets text,
  reps_or_time text,
  equipment text,
  intensity text,
  weight text,
  actual_sets text,
  actual_weight text,
  rest text,
  video_url text,
  notes text,
  order_num integer,
  created_at timestamp with time zone default now()
);

alter table public.athlete_block_exercises add column if not exists actual_sets text;
alter table public.athlete_block_exercises add column if not exists actual_weight text;

alter table public.athlete_blocks no force row level security;
alter table public.athlete_blocks disable row level security;
alter table public.athlete_block_exercises no force row level security;
alter table public.athlete_block_exercises disable row level security;

drop policy if exists "Development anon manage athlete blocks" on public.athlete_blocks;
create policy "Development anon manage athlete blocks"
on public.athlete_blocks
for all
to anon, authenticated
using (true)
with check (true);

drop policy if exists "Development anon manage athlete block exercises" on public.athlete_block_exercises;
create policy "Development anon manage athlete block exercises"
on public.athlete_block_exercises
for all
to anon, authenticated
using (true)
with check (true);

grant usage on schema public to anon, authenticated;
grant select, insert, update, delete on table public.athlete_blocks to anon, authenticated;
grant select, insert, update, delete on table public.athlete_block_exercises to anon, authenticated;
grant usage, select on all sequences in schema public to anon, authenticated;

-- 跑完後可以用這段確認 RLS 是否真的關閉：
select
  relname,
  relrowsecurity,
  relforcerowsecurity
from pg_class
where relname in ('athlete_blocks', 'athlete_block_exercises');
        """.strip(),
        language="sql",
    )


def block_label(row):
    if not isinstance(row, dict):
        return "未命名板塊"
    block_code = row.get("block_code")
    block_name = row.get("block_name")
    if has_value(block_code) and has_value(block_name):
        return f"{block_code} | {block_name}"
    return block_name or block_code or f"Block {row.get('id')}"


def render_labeled_value(label, value):
    st.markdown(f"**{label}：** {value or '-'}")


def render_event_input(key_prefix, current_value=""):
    current_value = current_value if has_value(current_value) else ""
    event_name = st.text_input(
        "賽事/事件",
        value=current_value,
        placeholder="輸入賽事或事件",
        key=f"{key_prefix}_event",
    )
    return event_name.strip()


def coach_page():
    st.title("Coach")
    st.caption("新增學員、查看學員，並直接替每位學員加入訓練板塊")

    athletes_df, fetch_error = fetch_athletes()
    if fetch_error:
        render_setup_help(fetch_error)
        return

    query_athlete_id = to_plain_int(get_query_param("assign_athlete"))
    if query_athlete_id and not athletes_df.empty:
        athlete_ids = set(athletes_df["id"].apply(to_plain_int).dropna().astype(int).tolist())
        if query_athlete_id in athlete_ids:
            st.session_state["selected_athlete_id"] = query_athlete_id

    selected_athlete_id = st.session_state.get("selected_athlete_id")
    if selected_athlete_id:
        render_selected_athlete_page(athletes_df, selected_athlete_id)
        return

    render_new_athlete_form()

    st.subheader("學員列表")
    if athletes_df.empty:
        st.info("目前尚未建立學員。")
    else:
        athlete_search = st.text_input(
            "搜尋學員",
            placeholder="輸入姓名、Email 或運動項目",
            key="coach_athlete_search",
        ).strip().lower()

        filtered_athletes_df = rank_athletes_by_search(athletes_df, athlete_search)

        if filtered_athletes_df.empty:
            st.info("找不到符合搜尋條件的學員。")
        else:
            render_clickable_athlete_list(filtered_athletes_df)


def render_new_athlete_form():
    with st.expander("新增學員", expanded=False):
        with st.container(border=True):
            with st.form("new_athlete_form", clear_on_submit=True):
                col1, col2 = st.columns(2)
                name = col1.text_input("姓名")
                email = col2.text_input("Email")
                sport = st.text_input("運動項目")
                level = ""
                submitted = st.form_submit_button("新增學員", use_container_width=True)

            if submitted:
                st.session_state.pop("last_temp_password", None)
                st.session_state.pop("last_temp_email", None)
                email = email.strip().lower()
                if not name or not email:
                    st.warning("請先輸入學員姓名和 Email。")
                elif athlete_email_exists(email):
                    st.error("這個 Email 已經存在，不能重複建立學員帳號。")
                elif not admin_supabase:
                    render_service_role_help()
                else:
                    try:
                        temp_password = generate_temp_password()
                        auth_user, auth_message = create_or_link_auth_user_for_athlete(
                            name,
                            email,
                            temp_password,
                        )
                        create_athlete(
                            name,
                            email,
                            sport,
                            level,
                            user_id=auth_user.id,
                            must_change_password=True,
                        )
                        st.session_state["last_temp_password"] = temp_password
                        st.session_state["last_temp_email"] = email
                        st.success(f"{auth_message} 已新增學員。")
                        st.rerun()
                    except Exception as exc:
                        render_auth_creation_error(exc)

            render_temp_password_card(
                "last_temp_password",
                "last_temp_email",
                "close_last_temp_password",
            )


def score_search_text(value, query, exact_weight, prefix_weight, token_prefix_weight, contains_weight):
    text = str(value or "").strip().lower()
    if not text or not query:
        return 0

    if text == query:
        return exact_weight

    score = 0
    if text.startswith(query):
        score = max(score, prefix_weight)

    normalized_tokens = (
        text.replace("@", " ")
        .replace(".", " ")
        .replace("_", " ")
        .replace("-", " ")
        .split()
    )
    if any(token.startswith(query) for token in normalized_tokens):
        score = max(score, token_prefix_weight)

    if query in text:
        score = max(score, contains_weight)

    return score


def rank_athletes_by_search(athletes_df, query):
    if not query:
        return athletes_df

    ranked_rows = []
    for index, athlete in athletes_df.iterrows():
        score = 0
        score += score_search_text(
            athlete.get("name"),
            query,
            exact_weight=240,
            prefix_weight=180,
            token_prefix_weight=140,
            contains_weight=100,
        )
        score += score_search_text(
            athlete.get("email"),
            query,
            exact_weight=220,
            prefix_weight=160,
            token_prefix_weight=120,
            contains_weight=90,
        )
        score += score_search_text(
            athlete.get("sport"),
            query,
            exact_weight=180,
            prefix_weight=130,
            token_prefix_weight=100,
            contains_weight=70,
        )

        if score > 0:
            ranked_rows.append((index, score))

    if not ranked_rows:
        return athletes_df.iloc[0:0]

    ranked_indexes = [index for index, _ in sorted(ranked_rows, key=lambda item: item[1], reverse=True)]
    return athletes_df.loc[ranked_indexes]


def render_clickable_athlete_list(athletes_df):
    for athlete in athletes_df.to_dict("records"):
        with st.container(border=True):
            st.markdown(f"### {athlete.get('name') or '未命名學員'}")

            info_cols = st.columns([2, 1])
            info_cols[0].markdown(f"**Email：** {athlete.get('email') or '-'}")
            info_cols[1].markdown(f"**運動項目：** {athlete.get('sport') or '-'}")

            action_cols = st.columns([3, 1])
            if action_cols[0].button("查看課表", key=f"open_athlete_{athlete['id']}", use_container_width=True):
                st.session_state["selected_athlete_id"] = athlete["id"]
                st.rerun()

            with action_cols[1]:
                with st.popover("☰", use_container_width=True):
                    if st.button("重設臨時密碼", key=f"reset_password_from_list_{athlete['id']}", use_container_width=True):
                        st.session_state["pending_reset_athlete_id"] = athlete["id"]
                    if st.button("刪除學員", key=f"ask_delete_athlete_{athlete['id']}", use_container_width=True):
                        st.session_state["pending_delete_athlete_id"] = athlete["id"]

            if st.session_state.get("pending_reset_athlete_id") == athlete["id"]:
                render_reset_password_confirmation(athlete)

            if st.session_state.get("pending_delete_athlete_id") == athlete["id"]:
                render_delete_athlete_confirmation(athlete)

            if st.session_state.get("last_reset_temp_athlete_id") == athlete["id"]:
                render_temp_password_card(
                    "last_reset_temp_password",
                    "last_reset_temp_email",
                    f"close_last_reset_temp_password_{athlete['id']}",
                    extra_keys=["last_reset_temp_athlete_id"],
                )


def reset_temp_password_from_list(athlete):
    if not admin_supabase:
        render_service_role_help()
        return

    try:
        temp_password, message = create_or_reset_athlete_temp_password(athlete)
        st.session_state["last_reset_temp_password"] = temp_password
        st.session_state["last_reset_temp_email"] = athlete.get("email")
        st.session_state["last_reset_temp_athlete_id"] = athlete["id"]
        st.success(message)
    except Exception as exc:
        st.error(f"重設失敗：{exc}")


def render_reset_password_confirmation(athlete):
    athlete_name = athlete.get("name") or athlete.get("email") or "這位學員"

    with st.container(border=True):
        st.warning(f"確認要重設 {athlete_name} 的臨時密碼嗎？重設後舊密碼會失效。")
        col1, col2 = st.columns(2)
        if col1.button("確認重設", key=f"confirm_reset_password_{athlete['id']}", use_container_width=True):
            st.session_state.pop("pending_reset_athlete_id", None)
            reset_temp_password_from_list(athlete)
        if col2.button("取消", key=f"cancel_reset_password_{athlete['id']}", use_container_width=True):
            st.session_state.pop("pending_reset_athlete_id", None)
            st.rerun()


def render_delete_athlete_confirmation(athlete):
    athlete_name = athlete.get("name") or athlete.get("email") or "這位學員"

    with st.container(border=True):
        st.warning(f"確認要刪除 {athlete_name} 嗎？刪除後會移除學員資料，並清掉已安排的課表板塊。")
        if has_value(athlete.get("user_id")) and not admin_supabase:
            st.info("目前沒有設定 service role key，所以只能刪除學員資料，無法同步刪除 Supabase Auth 帳號。")

        col1, col2 = st.columns(2)
        if col1.button("確認刪除", key=f"confirm_delete_athlete_{athlete['id']}", use_container_width=True):
            try:
                delete_message = delete_athlete(
                    athlete["id"],
                    athlete.get("user_id"),
                )
                if st.session_state.get("selected_athlete_id") == athlete["id"]:
                    st.session_state.pop("selected_athlete_id", None)
                st.session_state.pop("pending_delete_athlete_id", None)
                st.success(delete_message)
                st.rerun()
            except Exception as exc:
                st.error(f"刪除學員失敗：{exc}")
        if col2.button("取消", key=f"cancel_delete_athlete_{athlete['id']}", use_container_width=True):
            st.session_state.pop("pending_delete_athlete_id", None)
            st.rerun()


def render_selected_athlete_page(athletes_df, athlete_id):
    athlete_id = to_plain_int(athlete_id)
    selected_rows = athletes_df[athletes_df["id"].apply(to_plain_int) == athlete_id]
    if selected_rows.empty:
        st.session_state.pop("selected_athlete_id", None)
        clear_assignment_query_params()
        st.warning("找不到這位學員，已返回學員列表。")
        st.rerun()

    if st.button("返回學員列表", key="back_to_athlete_list"):
        st.session_state.pop("selected_athlete_id", None)
        clear_assignment_query_params()
        st.rerun()

    st.divider()
    selected_athlete = selected_rows.iloc[0].to_dict()
    render_athlete_program_section(selected_athlete)


def blocks_page():
    st.title("板塊")
    st.caption("新增訓練板塊，並查看目前所有板塊")

    blocks_df, blocks_error = fetch_blocks()
    if blocks_error:
        render_block_setup_help(blocks_error)
        return

    with st.expander("從 Excel 匯入板塊", expanded=False):
        uploaded_file = st.file_uploader(
            "上傳板塊 Excel 檔",
            type=["xlsx"],
            help="支援你目前的 LAB33 板塊模板格式。",
        )

        if uploaded_file:
            try:
                uploaded_file.seek(0)
                parsed_blocks = parse_block_workbook(uploaded_file)

                if not parsed_blocks:
                    st.warning("沒有讀到可匯入的模板。請確認每個工作表都有 LAB33 板塊模板格式。")
                    return

                importable_blocks, skipped_rows = split_importable_blocks(parsed_blocks, blocks_df)

                st.markdown("#### 匯入預覽")
                importable_summary_rows = []
                for parsed_block in importable_blocks:
                    block_code = parsed_block["sheet_name"]
                    block_name = parsed_block["block_name"]
                    exercise_count = sum(len(section["exercises"]) for section in parsed_block["sections"])
                    importable_summary_rows.append(
                        {
                            "匯入": True,
                            "將寫入 Block Code": block_code,
                            "將寫入顯示名稱": block_name,
                            "週期目標": parsed_block.get("goal") or "",
                            "訓練元素": parsed_block.get("training_element") or "",
                            "區段數": len(parsed_block["sections"]),
                            "動作數": exercise_count,
                        }
                    )
                st.success(f"已讀取 {len(parsed_blocks)} 個板塊模板，可匯入 {len(importable_blocks)} 個。")
                if importable_summary_rows:
                    selected_summary_df = st.data_editor(
                        pd.DataFrame(importable_summary_rows),
                        use_container_width=True,
                        hide_index=True,
                        key=f"import_block_selection_{uploaded_file.name}_{len(importable_blocks)}",
                        column_config={
                            "匯入": st.column_config.CheckboxColumn(
                                "匯入",
                                help="取消勾選就不匯入這張工作表。",
                                default=True,
                            )
                        },
                        disabled=[
                            "將寫入 Block Code",
                            "將寫入顯示名稱",
                            "週期目標",
                            "訓練元素",
                            "區段數",
                            "動作數",
                        ],
                    )
                else:
                    selected_summary_df = pd.DataFrame()

                if skipped_rows:
                    st.warning("以下板塊會跳過，不影響其他可匯入的工作表。")
                    st.dataframe(pd.DataFrame(skipped_rows), use_container_width=True, hide_index=True)

                selected_block_codes = set()
                if not selected_summary_df.empty:
                    selected_block_codes = set(
                        selected_summary_df[
                            selected_summary_df["匯入"]
                        ]["將寫入 Block Code"].tolist()
                    )
                selected_blocks = [
                    parsed_block for parsed_block in importable_blocks
                    if parsed_block["sheet_name"] in selected_block_codes
                ]

                for parsed_block in selected_blocks:
                    st.markdown(f"**{parsed_block['sheet_name']}｜{parsed_block['block_name']}**")
                    preview_rows = []
                    for section in parsed_block["sections"]:
                        for exercise in section["exercises"]:
                            preview_rows.append(
                                {
                                    "section": section["section_name"],
                                    **exercise,
                                }
                            )
                    if preview_rows:
                        st.dataframe(pd.DataFrame(preview_rows), use_container_width=True, hide_index=True)
                    else:
                        st.warning("這張工作表沒有讀到動作列。")

                with st.form("import_block_excel_form"):
                    import_description = st.text_area("描述", value="由 Excel 多工作表匯入")
                    import_submitted = st.form_submit_button(
                        f"確認匯入選取的 {len(selected_blocks)} 個板塊",
                        use_container_width=True,
                    )

                if import_submitted:
                    if not selected_blocks:
                        st.warning("目前沒有選取任何可匯入的新板塊。")
                        return

                    for parsed_block in selected_blocks:
                        block_code = parsed_block["sheet_name"]
                        create_block_from_excel(
                            parsed_block,
                            block_code,
                            import_description,
                        )
                    st.success(f"已從 Excel 匯入 {len(selected_blocks)} 個板塊。")
                    st.rerun()
            except Exception as exc:
                render_block_setup_help(exc)

    render_manual_block_template_creator()

    if blocks_df.empty:
        st.info("目前尚未建立板塊。")
    else:
        render_blocks_with_details(blocks_df)


def render_manual_block_template_creator():
    with st.expander("直接填寫模板", expanded=False):
        st.caption("依照 LAB33 板塊模板填入動作；每個有填動作名稱的列都會寫進資料庫。")
        version = manual_template_version()

        col1, col2 = st.columns(2)
        block_code = col1.text_input("Block Code", key=f"manual_block_code_{version}")
        block_name = col2.text_input("顯示名稱（留空會用 Block Code）", key=f"manual_block_name_{version}")
        goal = col1.text_area("週期目標", key=f"manual_block_goal_{version}", height=120)
        training_element = col2.text_area("訓練元素", key=f"manual_training_element_{version}", height=120)
        description = st.text_area("描述", key=f"manual_block_description_{version}", height=160)

        section_tables = {}
        column_config = {
            "exercise_name": st.column_config.TextColumn("動作"),
            "sets": st.column_config.TextColumn("組數"),
            "reps_or_time": st.column_config.TextColumn("次數/時間"),
            "equipment": st.column_config.TextColumn("工具"),
            "intensity": st.column_config.TextColumn("強度"),
            "weight": st.column_config.TextColumn("重量"),
            "rest": st.column_config.TextColumn("休息時間"),
            "video_url": st.column_config.TextColumn("影片連結"),
        }

        for section_name in MANUAL_BLOCK_SECTION_NAMES:
            st.markdown(f"### {section_name}")
            section_tables[section_name] = st.data_editor(
                empty_manual_template_rows(),
                key=f"manual_template_{section_name}_{version}",
                num_rows="dynamic",
                use_container_width=True,
                hide_index=True,
                column_order=MANUAL_TEMPLATE_COLUMNS,
                column_config=column_config,
            )

        if st.button("建立這個板塊模板", key=f"create_manual_template_block_{version}", use_container_width=True):
            display_name = block_name or block_code
            if not display_name:
                st.warning("請先輸入 Block Code 或顯示名稱。")
                return

            parsed_block = build_manual_template_block(
                display_name,
                goal,
                training_element,
                section_tables,
            )
            exercise_count = sum(len(section["exercises"]) for section in parsed_block["sections"])
            if exercise_count == 0:
                st.warning("請至少填入一個動作名稱。")
                return

            identity_errors = validate_new_block_identity(
                [
                    {
                        "block_code": block_code or display_name,
                        "block_name": display_name,
                    }
                ],
                fetch_blocks()[0],
            )
            if identity_errors:
                render_block_identity_errors(identity_errors)
                return

            try:
                create_block_from_excel(
                    parsed_block,
                    block_code or display_name,
                    description or "由手動模板建立",
                )
                st.success(f"已建立板塊模板，共寫入 {exercise_count} 個動作。")
                st.rerun()
            except Exception as exc:
                render_block_setup_help(exc)

        if st.button("清空模板內容", key=f"clear_manual_template_{version}", use_container_width=True):
            reset_manual_template()
            st.rerun()


def render_blocks_with_details(blocks_df):
    st.markdown("#### 板塊詳細內容")
    for block in blocks_df.to_dict("records"):
        with st.expander(f"{block.get('block_code') or block.get('id')}｜{block.get('block_name') or '未命名板塊'}"):
            render_labeled_value("目標", block.get("goal"))
            render_labeled_value("訓練元素", block.get("training_element"))
            col3, col4 = st.columns([5, 1])
            col3.write(f"描述：{block.get('description') or '-'}")
            if col4.button("刪除", key=f"ask_delete_block_{block['id']}", use_container_width=True):
                st.session_state["pending_delete_block_id"] = block["id"]

            if st.session_state.get("pending_delete_block_id") == block["id"]:
                render_delete_block_confirmation(block)

            render_block_detail_tables(block["id"])


def format_block_exercise_table(exercises_df):
    visible_columns = [
        "exercise_name",
        "sets",
        "reps_or_time",
        "equipment",
        "intensity",
        "weight",
        "rest",
        "video_url",
        "notes",
    ]
    column_labels = {
        "exercise_name": "動作",
        "sets": "組數",
        "reps_or_time": "次數/時間",
        "equipment": "工具",
        "intensity": "強度",
        "weight": "重量",
        "rest": "休息時間",
        "video_url": "影片連結",
        "notes": "備註",
    }
    available_columns = [
        column for column in visible_columns
        if column in exercises_df.columns
    ]
    return exercises_df[available_columns].rename(columns=column_labels)


def block_exercise_display_column_config():
    return {
        "影片連結": st.column_config.LinkColumn(
            "影片連結",
            display_text="觀看影片",
            help="點擊後會開啟影片連結",
        )
    }


def exercise_editor_column_config():
    return {
        "exercise_name": st.column_config.TextColumn("動作"),
        "sets": st.column_config.TextColumn("組數"),
        "reps_or_time": st.column_config.TextColumn("次數/時間"),
        "equipment": st.column_config.TextColumn("工具"),
        "intensity": st.column_config.TextColumn("強度"),
        "weight": st.column_config.TextColumn("重量"),
        "rest": st.column_config.TextColumn("休息時間"),
        "video_url": st.column_config.TextColumn("影片連結"),
        "notes": st.column_config.TextColumn("備註"),
    }


def assignment_exercises_from_template(block_id):
    return pd.DataFrame(template_exercise_rows_for_block(block_id))


def assignment_exercises_for_editing(assignment_id, block_id):
    exercises_df, error = fetch_athlete_block_exercises(assignment_id)
    if error:
        return pd.DataFrame(), error
    if exercises_df.empty:
        try:
            return assignment_exercises_from_template(block_id), None
        except Exception as exc:
            return pd.DataFrame(), exc
    return exercises_df, None


def ensure_report_columns(exercises_df):
    if exercises_df is None:
        return pd.DataFrame()
    exercises_df = exercises_df.copy()
    for column in ["actual_sets", "actual_weight"]:
        if column not in exercises_df.columns:
            exercises_df[column] = ""
    return exercises_df


def build_assignment_report_editor_rows(exercises_df):
    exercises_df = ensure_report_columns(exercises_df)
    if exercises_df.empty:
        return pd.DataFrame()

    for column in ["id", "section_name", "exercise_name", "sets", "weight"]:
        if column not in exercises_df.columns:
            exercises_df[column] = ""

    editor_rows = exercises_df[
        [
            "id",
            "section_name",
            "exercise_name",
            "sets",
            "weight",
            "actual_sets",
            "actual_weight",
        ]
    ].copy()
    return editor_rows.rename(
        columns={
            "section_name": "區段",
            "exercise_name": "動作",
            "sets": "安排組數",
            "weight": "安排重量",
            "actual_sets": "實際組數",
            "actual_weight": "實際重量",
        }
    )


def render_assignment_report_form(assignment_id, exercises_df, form_key_prefix="assignment_report"):
    exercises_df = ensure_report_columns(exercises_df)
    if exercises_df.empty:
        st.info("這個板塊目前沒有可回報的動作內容。")
        return
    if "id" not in exercises_df.columns or exercises_df["id"].isna().all():
        st.info("這個板塊還沒有建立可儲存回報的學員專屬動作內容。")
        return

    editor_rows = build_assignment_report_editor_rows(exercises_df)
    if editor_rows.empty:
        return

    st.markdown("#### 訓練回報")
    st.caption("填寫實際完成內容；這裡不會修改教練原本安排的組數或重量。")
    form_key = f"{form_key_prefix}_{to_plain_int(assignment_id) or assignment_id}"
    with st.form(form_key):
        edited_rows = st.data_editor(
            editor_rows.drop(columns=["id"]),
            hide_index=True,
            use_container_width=True,
            disabled=["區段", "動作", "安排組數", "安排重量"],
            column_config={
                "區段": st.column_config.TextColumn("區段"),
                "動作": st.column_config.TextColumn("動作"),
                "安排組數": st.column_config.TextColumn("安排組數"),
                "安排重量": st.column_config.TextColumn("安排重量"),
                "實際組數": st.column_config.TextColumn(
                    "實際組數",
                    help="填寫實際完成的組數",
                ),
                "實際重量": st.column_config.TextColumn(
                    "實際重量",
                    help="填寫實際使用的重量，例如 40kg 或 90lb",
                ),
            },
        )
        submitted = st.form_submit_button("儲存回報", use_container_width=True)

    if submitted:
        try:
            for index in range(len(edited_rows)):
                update_athlete_exercise_report(
                    editor_rows.iloc[index]["id"],
                    edited_rows.iloc[index].get("實際組數"),
                    edited_rows.iloc[index].get("實際重量"),
                )
            st.success("已儲存訓練回報。")
        except Exception as exc:
            render_block_assignment_setup_help(exc)


def render_assignment_detail_tables(
    assignment_id,
    block_id,
    empty_message="這個板塊目前沒有詳細內容。",
    show_report=False,
    report_key_prefix="assignment_report",
):
    exercises_df, error = fetch_athlete_block_exercises(assignment_id)
    if error:
        render_block_assignment_setup_help(error)
        return
    if exercises_df.empty:
        exercises_df = assignment_exercises_from_template(block_id)

    if exercises_df.empty:
        st.info(empty_message)
        return

    exercises_df = ensure_report_columns(exercises_df)
    section_names = exercises_df.sort_values(["section_order", "order_num"])["section_name"].dropna().unique()
    for section_name in section_names:
        st.markdown(f"**{section_name or '未命名區段'}**")
        section_df = exercises_df[exercises_df["section_name"] == section_name].copy()
        st.dataframe(
            format_block_exercise_table(section_df),
            use_container_width=True,
            hide_index=True,
            column_config=block_exercise_display_column_config(),
        )

    if show_report:
        render_assignment_report_form(assignment_id, exercises_df, report_key_prefix)


def assignment_report_label(assignment_row, blocks_by_id):
    block = blocks_by_id.get(to_plain_int(assignment_row.get("block_id")))
    block_text = block_label(block) if block else str(assignment_row.get("block_id") or "未命名板塊")
    parts = []
    date_text = date_range_label(assignment_row)
    if date_text:
        parts.append(date_text)
    parts.append(f"Week {assignment_row.get('week_num') or '-'} / Day {assignment_row.get('day_num') or '-'}")
    if has_value(assignment_row.get("training_category")):
        parts.append(str(assignment_row.get("training_category")))
    parts.append(block_text)
    return " | ".join(parts)


def render_student_training_report(athlete_blocks_df, blocks_df):
    if athlete_blocks_df.empty:
        return

    blocks_by_id = {
        to_plain_int(block.get("id")): block
        for block in blocks_df.to_dict("records")
    }
    assignment_rows = [
        row for row in athlete_blocks_df.to_dict("records")
        if to_plain_int(row.get("id")) is not None
    ]
    if not assignment_rows:
        return

    assignment_rows = sorted(
        assignment_rows,
        key=lambda row: (
            str(row.get("scheduled_date") or row.get("start_date") or ""),
            to_plain_int(row.get("week_num")) or 0,
            to_plain_int(row.get("day_num")) or 0,
            to_plain_int(row.get("id")) or 0,
        ),
    )
    assignment_by_id = {
        to_plain_int(row.get("id")): row for row in assignment_rows
    }

    with st.expander("訓練回報", expanded=False):
        st.caption("填寫實際完成內容；這裡不會修改教練原本安排的組數或重量。")
        selected_assignment_id = st.selectbox(
            "選擇要回報的課表",
            list(assignment_by_id.keys()),
            format_func=lambda assignment_id: assignment_report_label(
                assignment_by_id.get(assignment_id, {}),
                blocks_by_id,
            ),
            key="student_report_assignment",
        )

        exercises_df, error = fetch_athlete_block_exercises(selected_assignment_id)
        if error:
            render_block_assignment_setup_help(error)
            return
        if exercises_df.empty:
            st.info("這個板塊目前沒有可回報的動作內容。")
            return
        if "id" not in exercises_df.columns or exercises_df["id"].isna().all():
            st.info("這個板塊還沒有建立可儲存回報的學員專屬動作內容。")
            return

        for column in ["actual_sets", "actual_weight"]:
            if column not in exercises_df.columns:
                exercises_df[column] = ""

        editor_rows = exercises_df[
            [
                "id",
                "section_name",
                "exercise_name",
                "sets",
                "weight",
                "actual_sets",
                "actual_weight",
            ]
        ].copy()
        editor_rows = editor_rows.rename(
            columns={
                "section_name": "區段",
                "exercise_name": "動作",
                "sets": "安排組數",
                "weight": "安排重量",
                "actual_sets": "實際組數",
                "actual_weight": "實際重量",
            }
        )

        with st.form(f"student_training_report_form_{selected_assignment_id}"):
            edited_rows = st.data_editor(
                editor_rows.drop(columns=["id"]),
                hide_index=True,
                use_container_width=True,
                disabled=["區段", "動作", "安排組數", "安排重量"],
                column_config={
                    "區段": st.column_config.TextColumn("區段"),
                    "動作": st.column_config.TextColumn("動作"),
                    "安排組數": st.column_config.TextColumn("安排組數"),
                    "安排重量": st.column_config.TextColumn("安排重量"),
                    "實際組數": st.column_config.TextColumn(
                        "實際組數",
                        help="填寫實際完成的組數",
                    ),
                    "實際重量": st.column_config.TextColumn(
                        "實際重量",
                        help="填寫實際使用的重量，例如 40kg 或 90lb",
                    ),
                },
            )
            submitted = st.form_submit_button("儲存回報", use_container_width=True)

        if submitted:
            try:
                for index in range(len(edited_rows)):
                    update_athlete_exercise_report(
                        editor_rows.iloc[index]["id"],
                        edited_rows.iloc[index].get("實際組數"),
                        edited_rows.iloc[index].get("實際重量"),
                    )
                st.success("已儲存訓練回報。")
                st.rerun()
            except Exception as exc:
                render_block_assignment_setup_help(exc)


def render_block_detail_tables(block_id, empty_message="這個板塊目前沒有詳細內容。"):
    sections_df, sections_error = fetch_block_sections(block_id)
    exercises_df, exercises_error = fetch_block_exercises(block_id)

    if sections_error or exercises_error:
        render_block_setup_help(sections_error or exercises_error)
        return

    if sections_df.empty and exercises_df.empty:
        st.info(empty_message)
        return

    for section in sections_df.to_dict("records"):
        st.markdown(f"**{section.get('section_name') or '未命名區段'}**")
        if exercises_df.empty or "section_id" not in exercises_df.columns:
            st.caption("此區段尚未建立動作。")
            continue

        section_exercises = exercises_df[
            exercises_df["section_id"] == section["id"]
        ].copy()
        if section_exercises.empty:
            st.caption("此區段尚未建立動作。")
            continue

        st.dataframe(
            format_block_exercise_table(section_exercises),
            use_container_width=True,
            hide_index=True,
            column_config=block_exercise_display_column_config(),
        )


def render_delete_block_confirmation(block):
    block_name = block.get("block_name") or block.get("block_code") or f"Block {block.get('id')}"

    with st.container(border=True):
        st.warning(
            f"確認要刪除 {block_name} 嗎？刪除後會移除板塊詳細內容，"
            "也會從已安排給學員的課表中移除。"
        )
        col1, col2 = st.columns(2)
        if col1.button("確認刪除", key=f"confirm_delete_block_{block['id']}", use_container_width=True):
            try:
                delete_message = delete_block(block["id"])
                st.session_state.pop("pending_delete_block_id", None)
                st.success(delete_message)
                st.rerun()
            except Exception as exc:
                render_block_setup_help(exc)
        if col2.button("取消", key=f"cancel_delete_block_{block['id']}", use_container_width=True):
            st.session_state.pop("pending_delete_block_id", None)
            st.rerun()


def render_athlete_program_section(selected_athlete):
    st.subheader("學員課表板塊")

    with st.container(border=True):
        st.markdown(f"### {selected_athlete.get('name') or '未命名學員'}")
        col1, col2, col3 = st.columns(3)
        col1.write(f"Email：{selected_athlete.get('email') or '-'}")
        col2.write(f"運動項目：{selected_athlete.get('sport') or '-'}")
        col3.write(f"學員 ID：{selected_athlete.get('id') or '-'}")

    blocks_df, blocks_error = fetch_blocks()
    if blocks_error:
        st.error(f"讀取 blocks 失敗：{blocks_error}")
        st.info("請先確認 Supabase 已有 `blocks` 資料表，且 anon key 可以讀取。")
        return

    athlete_blocks_df, athlete_blocks_error = fetch_athlete_blocks(selected_athlete["id"])
    if athlete_blocks_error:
        render_block_assignment_setup_help(athlete_blocks_error)
        return

    render_week_calendar_assignment(selected_athlete, blocks_df, athlete_blocks_df)

    if athlete_blocks_df.empty:
        st.info("這位學員目前還沒有加入任何板塊。")
    else:
        render_schedule_table(
            athlete_blocks_df,
            blocks_df,
            editable=True,
            show_summary=False,
            show_period_view=False,
        )


def render_week_calendar_assignment(selected_athlete, blocks_df, athlete_blocks_df):
    with st.container(border=True):
        st.subheader("加入板塊")
        if blocks_df.empty:
            st.info("目前沒有任何 block。請先在板塊頁新增板塊。")
            return

        block_options = blocks_df.to_dict("records")
        block_by_id = {
            to_plain_int(block["id"]): block
            for block in block_options
        }
        block_ids = list(block_by_id.keys())

        selected_calendar_range = render_assignment_calendar(
            selected_athlete["id"],
            athlete_blocks_df,
            block_by_id,
            editable_detail=True,
        )
        if not selected_calendar_range:
            st.info("請先在上方行事曆點選一天，或拖曳選取日期區間，再安排當天課表。")
            return

        start_date, end_date = selected_calendar_range
        date_text = start_date.isoformat() if start_date == end_date else f"{start_date.isoformat()} ~ {end_date.isoformat()}"
        st.markdown(f"#### 安排 {date_text} 的課表")
        with st.form(f"assign_block_form_{selected_athlete['id']}", clear_on_submit=True):
            event_name = render_event_input(f"assign_block_{selected_athlete['id']}")
            date_col1, date_col2 = st.columns(2)
            form_start_date = date_col1.date_input("開始日期", value=start_date)
            form_end_date = date_col2.date_input("結束日期", value=end_date)
            cycle_goal = st.text_area("週期目標", height=70)
            col1, col2 = st.columns(2)
            week_num = col1.number_input("Week", min_value=1, value=1)
            day_num = col2.number_input("Day", min_value=1, value=form_start_date.isoweekday())
            training_category = st.selectbox("訓練分類", TRAINING_CATEGORIES)
            selected_block_id = st.selectbox(
                "選擇板塊",
                block_ids,
                format_func=lambda block_id: block_label(block_by_id[block_id]),
            )
            notes = st.text_area("備註", height=70)
            submitted = st.form_submit_button("加入到這位學員課表", use_container_width=True)

        if submitted:
            if form_end_date < form_start_date:
                st.warning("結束日期不能早於開始日期。")
                return
            try:
                assign_block_to_athlete(
                    selected_athlete["id"],
                    selected_block_id,
                    event_name,
                    cycle_goal,
                    form_start_date,
                    form_end_date,
                    week_num,
                    day_num,
                    training_category,
                    notes,
                )
                st.session_state.pop(f"assignment_calendar_selected_range_{selected_athlete['id']}", None)
                st.success("已將板塊加入這位學員的課表。")
                st.rerun()
            except Exception as exc:
                render_block_assignment_setup_help(exc)


def assignment_calendar_options():
    return {
        "initialView": "dayGridMonth",
        "locale": "zh-tw",
        "firstDay": 1,
        "selectable": True,
        "height": 620,
        "dayMaxEvents": True,
        "headerToolbar": {
            "left": "prev,next today",
            "center": "title",
            "right": "dayGridMonth,listMonth",
        },
        "buttonText": {
            "today": "今天",
            "month": "月",
            "list": "清單",
        },
    }


def assignment_calendar_custom_css():
    return """
    .fc .fc-button-primary {
        background: #3E424B;
        border-color: #3E424B;
        box-shadow: none;
    }
    .fc .fc-button-primary:hover,
    .fc .fc-button-primary:focus {
        background: #292d34;
        border-color: #292d34;
    }
    .fc-event {
        border: 2px solid #3E424B !important;
        border-radius: 8px;
        cursor: pointer;
        font-weight: 700;
        padding: 2px 6px;
    }
    .fc-event.selected-calendar-event,
    .fc-event.selected-calendar-event:hover {
        border-color: #f59e0b !important;
        box-shadow: 0 0 0 2px rgba(245, 158, 11, 0.45) !important;
    }
    .fc-event:active,
    .fc-event:focus,
    .fc-event:focus-visible,
    .fc-event:focus-within,
    a.fc-event:focus,
    a.fc-event:focus-visible {
        outline: none !important;
        box-shadow: 0 0 0 2px #f59e0b !important;
    }
    @media (max-width: 640px) {
        .fc .fc-toolbar {
            align-items: stretch;
            flex-direction: column;
            gap: 0.5rem;
        }
        .fc .fc-toolbar-title {
            font-size: 1.1rem;
        }
        .fc .fc-daygrid-day-frame {
            min-height: 72px;
        }
    }
    """


def render_calendar_assignment_detail_card(
    assignment_id,
    athlete_blocks_df,
    block_by_id,
    editable=False,
):
    if athlete_blocks_df is None or athlete_blocks_df.empty or "id" not in athlete_blocks_df.columns:
        return

    assignment_rows = athlete_blocks_df[
        athlete_blocks_df["id"].astype(str) == str(assignment_id)
    ]
    if assignment_rows.empty:
        return

    row = assignment_rows.iloc[0].to_dict()
    block = block_by_id.get(to_plain_int(row.get("block_id")))
    block_name = block_label(block) if block else str(row.get("block_id") or "未命名板塊")
    title_parts = [
        f"Week {row.get('week_num')} / Day {row.get('day_num')}",
        date_range_label(row),
        row.get("training_category") or "未分類",
        block_name,
    ]

    with st.container(border=True):
        st.markdown(f"### {block_name}")
        st.caption("｜".join([str(part) for part in title_parts if has_value(part)]))
        if has_value(row.get("event_name")):
            st.write(f"**賽事/事件：** {row.get('event_name')}")
        if has_value(row.get("cycle_goal")):
            st.info(f"週期目標：{row.get('cycle_goal')}")
        if has_value(row.get("notes")):
            st.info(f"教練備註：{row.get('notes')}")

        if not block:
            st.warning("找不到這個板塊的詳細資料，可能已被刪除。")
            return

        render_labeled_value("目標", block.get("goal"))
        render_labeled_value("訓練元素", block.get("training_element"))
        if block.get("description"):
            st.caption(f"描述：{block.get('description')}")

        if editable:
            render_assignment_edit_form(
                row,
                block,
                form_key_prefix="calendar_edit_athlete_block_content",
            )
        else:
            render_assignment_detail_tables(
                row["id"],
                block["id"],
                empty_message="這個板塊目前沒有建立詳細動作內容。",
            )


def render_assignment_calendar(athlete_id, athlete_blocks_df, block_by_id, editable_detail=False):
    st.markdown("#### 課表行事曆")
    selection_key = f"assignment_calendar_selected_range_{athlete_id}"
    selected_event_key = f"assignment_calendar_selected_event_{athlete_id}"
    component_key = f"assignment_calendar_{athlete_id}"

    selected_event_id = sync_calendar_event_selection_from_state(
        component_key,
        selected_event_key,
    )
    events = build_assignment_calendar_events(
        athlete_blocks_df,
        block_by_id,
        selected_event_id,
    )

    calendar_range = None
    stored_range = st.session_state.get(selection_key)
    if stored_range:
        try:
            calendar_range = (
                pd.to_datetime(stored_range[0]).date(),
                pd.to_datetime(stored_range[1]).date(),
            )
        except Exception:
            st.session_state.pop(selection_key, None)

    if st_calendar is None:
        selected_date_text = ""
        if calendar_range and calendar_range[0] == calendar_range[1]:
            selected_date_text = calendar_range[0].isoformat()
            st.success(f"已選取 {selected_date_text}，可以直接在下方安排這一天的課表。")
        st.warning("目前缺少 streamlit-calendar，行事曆只能查看，不能直接點日期安排課表。")
        render_frontend_schedule_calendar(
            events,
            athlete_blocks_df,
            block_by_id,
            key_prefix=f"coach_schedule_{athlete_id}",
            enable_date_selection=False,
            selected_date=selected_date_text,
            athlete_id=athlete_id,
        )
        return calendar_range

    calendar_state = st_calendar(
        events=events,
        options=assignment_calendar_options(),
        custom_css=assignment_calendar_custom_css(),
        callbacks=["dateClick", "select", "eventClick"],
        key=component_key,
    )

    selected_event_id = sync_calendar_event_selection_from_state(
        calendar_state,
        selected_event_key,
    )

    selected_range = extract_calendar_selection(calendar_state)
    if selected_range:
        calendar_range = selected_range
        st.session_state[selection_key] = (
            calendar_range[0].isoformat(),
            calendar_range[1].isoformat(),
        )

    selected_date_text = ""
    if calendar_range and calendar_range[0] == calendar_range[1]:
        selected_date_text = calendar_range[0].isoformat()
        st.success(f"已選取 {selected_date_text}，可以直接在下方安排這一天的課表。")

    if selected_event_id:
        render_calendar_assignment_detail_card(
            selected_event_id,
            athlete_blocks_df,
            block_by_id,
            editable=editable_detail,
        )

    return calendar_range


@st.fragment
def render_student_schedule_calendar(athlete_id, athlete_blocks_df, blocks_df):
    st.markdown("#### 課表行事曆")
    block_by_id = {
        to_plain_int(block["id"]): block
        for block in blocks_df.to_dict("records")
    }
    events = build_assignment_calendar_events(
        athlete_blocks_df,
        block_by_id,
    )
    if not events:
        st.info("目前沒有可顯示在行事曆上的課表日期。")
        return

    render_frontend_schedule_calendar(
        events,
        athlete_blocks_df,
        block_by_id,
        key_prefix=f"student_schedule_{athlete_id}",
    )


def render_frontend_schedule_calendar(
    events,
    athlete_blocks_df,
    block_by_id,
    key_prefix,
    enable_date_selection=False,
    selected_date="",
    athlete_id=None,
):
    payload = {
        "events": events,
        "details": build_assignment_frontend_details(athlete_blocks_df, block_by_id),
        "enableDateSelection": enable_date_selection,
        "selectedDate": selected_date,
        "athleteId": str(athlete_id or ""),
    }
    payload_json = json.dumps(payload, ensure_ascii=False).replace("</", "<\\/")
    html = f"""
    <div id="{key_prefix}" class="lab33-calendar-root">
      <div class="lab33-cal-toolbar">
        <div class="lab33-cal-nav">
          <button type="button" data-action="prev">‹</button>
          <button type="button" data-action="today">今天</button>
          <button type="button" data-action="next">›</button>
        </div>
        <div class="lab33-cal-title"></div>
        <div class="lab33-cal-tabs">
          <button type="button" data-view="month" class="active">月</button>
          <button type="button" data-view="list">清單</button>
        </div>
      </div>
      <div class="lab33-cal-month"></div>
      <div class="lab33-cal-list" hidden></div>
      <div class="lab33-cal-detail" hidden></div>
    </div>
    <script>
      (() => {{
        const payload = {payload_json};
        const root = document.getElementById("{key_prefix}");
        const events = payload.events || [];
        const details = payload.details || {{}};
        const enableDateSelection = Boolean(payload.enableDateSelection);
        const selectedDate = payload.selectedDate || "";
        const athleteId = payload.athleteId || "";
        const weekdayLabels = ["週一", "週二", "週三", "週四", "週五", "週六", "週日"];
        const state = {{
          view: "month",
          selectedId: null,
          currentMonth: initialMonth(events),
        }};

        function parseIso(value) {{
          if (!value) return null;
          const parts = String(value).slice(0, 10).split("-").map(Number);
          if (parts.length !== 3 || parts.some(Number.isNaN)) return null;
          return new Date(parts[0], parts[1] - 1, parts[2]);
        }}

        function isoDate(date) {{
          const y = date.getFullYear();
          const m = String(date.getMonth() + 1).padStart(2, "0");
          const d = String(date.getDate()).padStart(2, "0");
          return `${{y}}-${{m}}-${{d}}`;
        }}

        function addDays(date, days) {{
          const next = new Date(date);
          next.setDate(next.getDate() + days);
          return next;
        }}

        function initialMonth() {{
          const today = new Date();
          return new Date(today.getFullYear(), today.getMonth(), 1);
        }}

        function eventDates(event) {{
          const start = parseIso(event.start);
          if (!start) return [];
          const end = event.end ? parseIso(event.end) : null;
          const exclusiveEnd = end && end > start ? end : addDays(start, 1);
          const dates = [];
          for (let day = new Date(start); day < exclusiveEnd; day = addDays(day, 1)) {{
            dates.push(isoDate(day));
          }}
          return dates;
        }}

        function eventsByDate() {{
          const grouped = {{}};
          events.forEach((event) => {{
            eventDates(event).forEach((dateKey) => {{
              grouped[dateKey] ||= [];
              grouped[dateKey].push(event);
            }});
          }});
          return grouped;
        }}

        function render() {{
          root.querySelector(".lab33-cal-title").textContent =
            `${{state.currentMonth.getFullYear()}} 年 ${{state.currentMonth.getMonth() + 1}} 月`;
          root.querySelectorAll("[data-view]").forEach((button) => {{
            button.classList.toggle("active", button.dataset.view === state.view);
          }});
          root.querySelector(".lab33-cal-month").hidden = state.view !== "month";
          root.querySelector(".lab33-cal-list").hidden = state.view !== "list";
          renderMonth();
          renderList();
          renderDetail();
        }}

        function renderMonth() {{
          const grouped = eventsByDate();
          const monthNode = root.querySelector(".lab33-cal-month");
          const firstDay = new Date(state.currentMonth);
          const offset = (firstDay.getDay() + 6) % 7;
          const gridStart = addDays(firstDay, -offset);
          let html = `<div class="lab33-cal-grid lab33-cal-weekdays">`;
          weekdayLabels.forEach((label) => html += `<div>${{label}}</div>`);
          html += `</div><div class="lab33-cal-grid">`;
          for (let i = 0; i < 42; i += 1) {{
            const day = addDays(gridStart, i);
            const dateKey = isoDate(day);
            const isCurrentMonth = day.getMonth() === state.currentMonth.getMonth();
            const isToday = dateKey === isoDate(new Date());
            const isSelectedDate = selectedDate === dateKey;
            html += `<div class="lab33-cal-cell ${{isCurrentMonth ? "" : "muted"}} ${{isToday ? "today" : ""}} ${{isSelectedDate ? "selected-date" : ""}}" data-date="${{dateKey}}">`;
            html += `<div class="lab33-cal-day">${{day.getDate()}}日</div>`;
            (grouped[dateKey] || []).forEach((event) => {{
              html += eventButton(event);
            }});
            html += `</div>`;
          }}
          html += `</div>`;
          monthNode.innerHTML = html;
        }}

        function renderList() {{
          const listNode = root.querySelector(".lab33-cal-list");
          const sortedEvents = [...events].sort((a, b) => String(a.start).localeCompare(String(b.start)));
          if (!sortedEvents.length) {{
            listNode.innerHTML = `<div class="lab33-empty">目前沒有課表。</div>`;
            return;
          }}
          listNode.innerHTML = sortedEvents.map((event) => {{
            const detail = details[String(event.id)] || {{}};
            const meta = [
              detail.date_range,
              detail.category,
              detail.block_label,
            ].filter(Boolean).join("｜");
            return `<div class="lab33-list-row">
              <div><strong>${{escapeHtml(event.title || "")}}</strong><div class="lab33-muted">${{escapeHtml(meta)}}</div></div>
              ${{eventButton(event)}}
            </div>`;
          }}).join("");
        }}

        function eventButton(event) {{
          const id = String(event.id);
          const isSelected = state.selectedId === id;
          const title = escapeHtml(event.title || "課表");
          return `<button type="button" class="lab33-event ${{isSelected ? "selected" : ""}}" data-event-id="${{escapeHtml(id)}}">${{title}}</button>`;
        }}

        function renderDetail() {{
          const detailNode = root.querySelector(".lab33-cal-detail");
          if (!state.selectedId) {{
            detailNode.hidden = true;
            detailNode.innerHTML = "";
            return;
          }}
          const detail = details[String(state.selectedId)];
          if (!detail) {{
            detailNode.hidden = true;
            detailNode.innerHTML = "";
            return;
          }}
          detailNode.hidden = false;
          detailNode.innerHTML = `
            <div class="lab33-detail-card">
              <h3>${{escapeHtml(detail.block_label || "課表")}}</h3>
              <div class="lab33-muted">${{escapeHtml(detail.meta || "")}}</div>
              ${{detail.event_name || detail.date_range ? `<p class="lab33-muted">賽事/事件：${{escapeHtml(detail.event_name || "-")}}｜日期：${{escapeHtml(detail.date_range || "-")}}</p>` : ""}}
              ${{detail.cycle_goal ? `<div class="lab33-goal"><strong>週期目標：</strong>${{escapeHtml(detail.cycle_goal)}}</div>` : ""}}
              ${{detail.goal ? `<p><strong>目標：</strong>${{escapeHtml(detail.goal)}}</p>` : ""}}
              ${{detail.training_element ? `<p><strong>訓練元素：</strong>${{escapeHtml(detail.training_element)}}</p>` : ""}}
              ${{detail.description ? `<p class="lab33-muted">描述：${{escapeHtml(detail.description)}}</p>` : ""}}
              ${{renderSections(detail.sections || [])}}
            </div>`;
        }}

        function renderSections(sections) {{
          if (!sections.length) return `<div class="lab33-empty">這個板塊目前沒有詳細動作內容。</div>`;
          return sections.map((section) => `
            <section class="lab33-section">
              <h4>${{escapeHtml(section.name || "未命名區段")}}</h4>
              <div class="lab33-table-wrap">
                <table>
                  <thead>
                    <tr>
                      <th>動作</th><th>組數</th><th>次數/時間</th><th>工具</th><th>強度</th><th>重量</th><th>休息時間</th><th>影片連結</th><th>備註</th>
                    </tr>
                  </thead>
                  <tbody>
                    ${{(section.rows || []).map((row) => `
                      <tr>
                        <td>${{escapeHtml(row.exercise_name)}}</td>
                        <td>${{escapeHtml(row.sets)}}</td>
                        <td>${{escapeHtml(row.reps_or_time)}}</td>
                        <td>${{escapeHtml(row.equipment)}}</td>
                        <td>${{escapeHtml(row.intensity)}}</td>
                        <td>${{escapeHtml(row.weight)}}</td>
                        <td>${{escapeHtml(row.rest)}}</td>
                        <td>${{row.video_url ? `<a href="${{escapeAttr(row.video_url)}}" target="_blank" rel="noopener noreferrer">觀看影片</a>` : ""}}</td>
                        <td>${{escapeHtml(row.notes)}}</td>
                      </tr>
                    `).join("")}}
                  </tbody>
                </table>
              </div>
            </section>`).join("");
        }}

        function escapeHtml(value) {{
          return String(value ?? "")
            .replaceAll("&", "&amp;")
            .replaceAll("<", "&lt;")
            .replaceAll(">", "&gt;")
            .replaceAll('"', "&quot;")
            .replaceAll("'", "&#039;");
        }}

        function escapeAttr(value) {{
          return escapeHtml(value).replaceAll("`", "&#096;");
        }}

        function selectDate(dateKey) {{
          if (!enableDateSelection || !athleteId || !dateKey) return;
          let baseUrl = "";
          try {{
            baseUrl = window.parent.location.href;
          }} catch (error) {{
            baseUrl = document.referrer || window.location.href;
          }}

          const parentUrl = new URL(baseUrl);
          parentUrl.searchParams.set("assign_athlete", athleteId);
          parentUrl.searchParams.set("assign_date", dateKey);
          parentUrl.searchParams.set("assign_ts", String(Date.now()));

          const targetUrl = parentUrl.toString();
          try {{
            window.parent.location.assign(targetUrl);
          }} catch (error) {{
            window.open(targetUrl, "_parent");
          }}
        }}

        root.addEventListener("click", (event) => {{
          const eventButtonNode = event.target.closest("[data-event-id]");
          if (eventButtonNode) {{
            const id = eventButtonNode.dataset.eventId;
            state.selectedId = state.selectedId === id ? null : id;
            render();
            return;
          }}
          const actionButton = event.target.closest("[data-action]");
          if (actionButton) {{
            const action = actionButton.dataset.action;
            if (action === "prev") state.currentMonth.setMonth(state.currentMonth.getMonth() - 1);
            if (action === "next") state.currentMonth.setMonth(state.currentMonth.getMonth() + 1);
            if (action === "today") state.currentMonth = new Date(new Date().getFullYear(), new Date().getMonth(), 1);
            render();
            return;
          }}
          const viewButton = event.target.closest("[data-view]");
          if (viewButton) {{
            state.view = viewButton.dataset.view;
            render();
            return;
          }}
          const dateCell = event.target.closest("[data-date]");
          if (dateCell && root.contains(dateCell)) {{
            selectDate(dateCell.dataset.date);
          }}
        }});

        render();
      }})();
    </script>
    <style>
      .lab33-calendar-root {{
        color: #30333f;
        font-family: "Source Sans Pro", sans-serif;
      }}
      .lab33-cal-toolbar {{
        align-items: center;
        display: flex;
        gap: 0.75rem;
        justify-content: space-between;
        margin-bottom: 0.75rem;
      }}
      .lab33-cal-title {{
        font-size: 1.15rem;
        font-weight: 800;
      }}
      .lab33-cal-nav,
      .lab33-cal-tabs {{
        display: flex;
        gap: 0.35rem;
      }}
      .lab33-cal-toolbar button {{
        background: #fff;
        border: 1px solid #d7d9df;
        border-radius: 10px;
        color: #30333f;
        cursor: pointer;
        font-weight: 700;
        padding: 0.45rem 0.7rem;
      }}
      .lab33-cal-tabs button.active {{
        background: #3E424B;
        border-color: #3E424B;
        color: #fff;
      }}
      .lab33-cal-grid {{
        display: grid;
        grid-template-columns: repeat(7, minmax(0, 1fr));
      }}
      .lab33-cal-weekdays > div {{
        border: 1px solid #e2e4e8;
        border-bottom: 0;
        font-weight: 800;
        padding: 0.45rem;
        text-align: center;
      }}
      .lab33-cal-cell {{
        border: 1px solid #e2e4e8;
        min-height: 118px;
        padding: 0.35rem;
      }}
      .lab33-cal-cell.muted {{
        color: #b7bbc4;
      }}
      .lab33-cal-cell.today {{
        background: #fff9df;
      }}
      .lab33-cal-cell.selected-date {{
        background: #e8f6fb;
        box-shadow: inset 0 0 0 2px #9bd3e4;
      }}
      .lab33-cal-cell[data-date] {{
        cursor: pointer;
      }}
      .lab33-cal-day {{
        font-size: 0.95rem;
        font-weight: 800;
        margin-bottom: 0.35rem;
        text-align: right;
      }}
      .lab33-event {{
        background: #3E424B;
        border: 2px solid #3E424B;
        border-radius: 9px;
        color: #fff;
        cursor: pointer;
        display: block;
        font-size: 0.78rem;
        font-weight: 800;
        margin: 0.22rem 0;
        max-width: 100%;
        overflow: hidden;
        padding: 0.22rem 0.4rem;
        text-align: left;
        text-overflow: ellipsis;
        white-space: nowrap;
      }}
      .lab33-event.selected {{
        background: #24272D;
        border-color: #f59e0b;
        box-shadow: 0 0 0 2px rgba(245, 158, 11, 0.55);
      }}
      .lab33-list-row {{
        align-items: center;
        border: 1px solid #e2e4e8;
        border-radius: 12px;
        display: flex;
        gap: 1rem;
        justify-content: space-between;
        margin-bottom: 0.5rem;
        padding: 0.75rem;
      }}
      .lab33-list-row .lab33-event {{
        flex: 0 0 auto;
        min-width: 120px;
      }}
      .lab33-detail-card {{
        border: 1px solid #d7d9df;
        border-radius: 14px;
        margin-top: 1rem;
        padding: 1rem;
      }}
      .lab33-detail-card h3 {{
        font-size: 1.45rem;
        margin: 0 0 0.45rem;
      }}
      .lab33-muted {{
        color: #7f8490;
      }}
      .lab33-goal {{
        background: #e9f3ff;
        border-radius: 10px;
        color: #0f4c81;
        margin: 0.75rem 0;
        padding: 0.75rem;
      }}
      .lab33-section h4 {{
        margin: 1rem 0 0.45rem;
      }}
      .lab33-table-wrap {{
        overflow-x: auto;
      }}
      .lab33-table-wrap table {{
        border-collapse: collapse;
        min-width: 920px;
        width: 100%;
      }}
      .lab33-table-wrap th,
      .lab33-table-wrap td {{
        border: 1px solid #e2e4e8;
        padding: 0.5rem;
        text-align: left;
        vertical-align: top;
      }}
      .lab33-table-wrap th {{
        background: #f6f7f9;
        color: #7f8490;
        font-weight: 700;
      }}
      .lab33-table-wrap a {{
        color: #0068c9;
        font-weight: 800;
      }}
      .lab33-empty {{
        background: #e9f3ff;
        border-radius: 10px;
        color: #0f4c81;
        padding: 0.8rem;
      }}
      @media (max-width: 640px) {{
        .lab33-cal-toolbar,
        .lab33-list-row {{
          align-items: stretch;
          flex-direction: column;
        }}
        .lab33-cal-title {{
          text-align: center;
        }}
        .lab33-cal-cell {{
          min-height: 86px;
          padding: 0.25rem;
        }}
        .lab33-cal-weekdays > div {{
          font-size: 0.8rem;
          padding: 0.3rem 0.1rem;
        }}
        .lab33-event {{
          font-size: 0.68rem;
          padding: 0.16rem 0.28rem;
        }}
      }}
    </style>
    """
    components.html(html, height=1180, scrolling=True)


def build_assignment_frontend_details(athlete_blocks_df, block_by_id):
    details = {}
    if athlete_blocks_df is None or athlete_blocks_df.empty:
        return details

    for row in athlete_blocks_df.to_dict("records"):
        assignment_id = str(row.get("id"))
        block = block_by_id.get(to_plain_int(row.get("block_id"))) or {}
        block_id = block.get("id") or row.get("block_id")
        block_name = block_label(block) if block else str(row.get("block_id") or "未命名板塊")
        exercises_df, error = fetch_athlete_block_exercises(row.get("id"))
        if error or exercises_df.empty:
            exercises_df = assignment_exercises_from_template(block_id)
        details[assignment_id] = {
            "block_label": block_name,
            "meta": "｜".join([
                f"Week {row.get('week_num')} / Day {row.get('day_num')}",
                date_range_label(row),
                row.get("training_category") or "未分類",
                block_name,
            ]),
            "event_name": safe_frontend_text(row.get("event_name")),
            "date_range": safe_frontend_text(date_range_label(row)),
            "cycle_goal": safe_frontend_text(row.get("cycle_goal")),
            "goal": safe_frontend_text(block.get("goal")),
            "training_element": safe_frontend_text(block.get("training_element")),
            "description": safe_frontend_text(block.get("description")),
            "sections": build_frontend_exercise_sections(exercises_df),
        }
    return details


def build_frontend_exercise_sections(exercises_df):
    if exercises_df is None or exercises_df.empty:
        return []

    exercises_df = exercises_df.copy()
    for column, default in {
        "section_name": "未命名區段",
        "section_order": 0,
        "order_num": 0,
    }.items():
        if column not in exercises_df.columns:
            exercises_df[column] = default

    sections = []
    sort_columns = [
        column for column in ["section_order", "order_num"]
        if column in exercises_df.columns
    ]
    if sort_columns:
        exercises_df = exercises_df.sort_values(sort_columns)

    for section_name, section_df in exercises_df.groupby("section_name", dropna=False):
        rows = []
        for exercise in section_df.to_dict("records"):
            rows.append({
                "exercise_name": safe_frontend_text(exercise.get("exercise_name")),
                "sets": safe_frontend_text(exercise.get("sets")),
                "reps_or_time": safe_frontend_text(exercise.get("reps_or_time")),
                "equipment": safe_frontend_text(exercise.get("equipment")),
                "intensity": safe_frontend_text(exercise.get("intensity")),
                "weight": safe_frontend_text(exercise.get("weight")),
                "rest": safe_frontend_text(exercise.get("rest")),
                "video_url": safe_frontend_text(exercise.get("video_url")),
                "notes": safe_frontend_text(exercise.get("notes")),
            })
        sections.append({
            "name": safe_frontend_text(section_name or "未命名區段"),
            "rows": rows,
        })
    return sections


def safe_frontend_text(value):
    return str(value) if has_value(value) else ""


def render_selected_student_assignment(selected_assignment_id, athlete_blocks_df, block_by_id):
    assignment_rows = athlete_blocks_df[
        athlete_blocks_df["id"].astype(str) == str(selected_assignment_id)
    ]
    if assignment_rows.empty:
        st.warning("找不到這筆課表內容，可能已被教練移除。")
        return

    row = assignment_rows.iloc[0].to_dict()
    block = block_by_id.get(to_plain_int(row.get("block_id")))
    block_name = block_label(block) if block else str(row.get("block_id") or "未命名板塊")
    title_parts = [
        f"Week {row.get('week_num')} / Day {row.get('day_num')}",
        date_range_label(row),
        row.get("training_category") or "未分類",
        block_name,
    ]
    title = "｜".join([str(part) for part in title_parts if has_value(part)])

    with st.container(border=True):
        st.markdown(f"### {block_name}")
        st.caption(title)
        if has_value(row.get("event_name")):
            st.write(f"**賽事/事件：** {row.get('event_name')}")
        if has_value(row.get("cycle_goal")):
            st.info(f"週期目標：{row.get('cycle_goal')}")
        if has_value(row.get("notes")):
            st.info(f"教練備註：{row.get('notes')}")

        if not block:
            st.warning("找不到這個板塊的詳細資料，可能已被刪除。")
            return

        render_labeled_value("目標", block.get("goal"))
        render_labeled_value("訓練元素", block.get("training_element"))
        if block.get("description"):
            st.caption(f"描述：{block.get('description')}")

        render_assignment_detail_tables(
            row["id"],
            block["id"],
            empty_message="這個板塊目前沒有建立詳細動作內容。",
            show_report=True,
            report_key_prefix="student_calendar_report",
        )


def build_assignment_calendar_events(athlete_blocks_df, block_by_id, selected_assignment_id=""):
    if athlete_blocks_df is None or athlete_blocks_df.empty:
        return []

    calendar_event_color = "#3E424B"
    events = []
    for row in athlete_blocks_df.to_dict("records"):
        start = calendar_date_string(row.get("start_date") or row.get("scheduled_date"))
        if not start:
            continue

        block = block_by_id.get(to_plain_int(row.get("block_id")))
        block_name = block_label(block) if block else str(row.get("block_id") or "未命名板塊")
        category = row.get("training_category") or "未分類"
        title = f"W{row.get('week_num')}D{row.get('day_num')}｜{category}｜{block_name}"
        event_color = calendar_event_color
        is_selected = str(row.get("id")) == str(selected_assignment_id)
        event = {
            "id": str(row.get("id")),
            "title": title,
            "start": start,
            "allDay": True,
            "backgroundColor": event_color,
            "borderColor": "#f59e0b" if is_selected else event_color,
            "textColor": "#ffffff",
            "classNames": ["selected-calendar-event"] if is_selected else [],
            "extendedProps": {
                "assignment_id": str(row.get("id")),
                "block": block_name,
                "category": category,
                "event_name": row.get("event_name") or "",
                "date_range": date_range_label(row),
            },
        }
        end = calendar_exclusive_end_string(row.get("end_date"), start)
        if end:
            event["end"] = end
        events.append(event)
    return events


def calendar_date_string(value):
    if not has_value(value):
        return ""
    try:
        return pd.to_datetime(value).date().isoformat()
    except Exception:
        return ""


def calendar_exclusive_end_string(end_value, start_iso):
    end_iso = calendar_date_string(end_value)
    if not end_iso or end_iso == start_iso:
        return ""
    try:
        return (pd.to_datetime(end_iso).date() + timedelta(days=1)).isoformat()
    except Exception:
        return end_iso


def extract_calendar_event(calendar_state):
    if not isinstance(calendar_state, dict):
        return ""
    event_click = calendar_state.get("eventClick") or {}
    event = event_click.get("event") if isinstance(event_click, dict) else {}
    if not isinstance(event, dict):
        return ""
    title = event.get("title")
    props = event.get("extendedProps") or {}
    detail_parts = [title] if title else []
    if props.get("event_name"):
        detail_parts.append(f"賽事/事件：{props.get('event_name')}")
    if props.get("date_range"):
        detail_parts.append(f"日期：{props.get('date_range')}")
    return "｜".join(detail_parts)


def extract_calendar_event_id(calendar_state):
    if not isinstance(calendar_state, dict):
        return ""
    event_click = calendar_state.get("eventClick") or {}
    event = event_click.get("event") if isinstance(event_click, dict) else {}
    if not isinstance(event, dict):
        return ""

    props = event.get("extendedProps") or {}
    event_def = event.get("_def") or {}
    event_id = (
        event.get("id")
        or props.get("assignment_id")
        or event_def.get("publicId")
    )
    return str(event_id) if has_value(event_id) else ""


def sync_calendar_event_selection_from_state(calendar_source, selection_key):
    if isinstance(calendar_source, str):
        calendar_state = st.session_state.get(calendar_source)
    else:
        calendar_state = calendar_source

    clicked_event_id = extract_calendar_event_id(calendar_state)
    if not clicked_event_id:
        return st.session_state.get(selection_key, "")

    event_click = calendar_state.get("eventClick") if isinstance(calendar_state, dict) else {}
    click_signature = json.dumps(event_click or {}, sort_keys=True, default=str)
    processed_key = f"{selection_key}_last_click_signature"
    if st.session_state.get(processed_key) == click_signature:
        return st.session_state.get(selection_key, "")

    selected_event_id = toggle_calendar_assignment_selection(selection_key, clicked_event_id)
    st.session_state[processed_key] = click_signature
    return selected_event_id


def toggle_calendar_assignment_selection(selection_key, clicked_event_id):
    current_id = st.session_state.get(selection_key)
    if not clicked_event_id:
        return current_id
    if str(current_id) == str(clicked_event_id):
        st.session_state.pop(selection_key, None)
        return ""
    st.session_state[selection_key] = clicked_event_id
    return clicked_event_id


def extract_calendar_selection(calendar_state):
    if not isinstance(calendar_state, dict):
        return None

    selected_payload = None
    # A single date click can leave a previous range selection in the payload.
    # Prefer the most recent click so the highlighted day and form date match.
    for key in ("dateClick", "select"):
        payload = calendar_state.get(key)
        if isinstance(payload, dict):
            selected_payload = payload
            break
    if not selected_payload:
        return None

    start_value = (
        selected_payload.get("startStr")
        or selected_payload.get("dateStr")
        or selected_payload.get("start")
        or selected_payload.get("date")
    )
    end_value = selected_payload.get("endStr") or selected_payload.get("end") or start_value
    start_date = parse_calendar_selection_date(start_value)
    end_date = parse_calendar_selection_date(end_value)
    if not start_date:
        return None
    if not end_date:
        end_date = start_date

    if end_date > start_date and (selected_payload.get("endStr") or selected_payload.get("end")):
        end_date = end_date - timedelta(days=1)
    if end_date < start_date:
        end_date = start_date
    return start_date, end_date


def parse_calendar_selection_date(value):
    if not has_value(value):
        return None
    if isinstance(value, str):
        raw_value = value.strip()
        if "T" not in raw_value and len(raw_value) >= 10:
            date_part = raw_value[:10]
            try:
                return date.fromisoformat(date_part)
            except ValueError:
                pass
    try:
        parsed_value = pd.to_datetime(value)
        if getattr(parsed_value, "tzinfo", None) is not None:
            parsed_value = parsed_value.tz_convert("Asia/Taipei")
        return parsed_value.date()
    except Exception:
        return None


def render_schedule_table(
    athlete_blocks_df,
    blocks_df,
    show_block_details=False,
    editable=False,
    show_summary=True,
    show_period_view=True,
    show_report=False,
):
    display_df = athlete_blocks_df.copy()
    block_records = blocks_df.to_dict("records")
    blocks_by_id = {
        row["id"]: row
        for row in block_records
    }
    block_names = {
        row["id"]: block_label(row)
        for row in block_records
    }
    display_df["block"] = display_df["block_id"].map(block_names).fillna(
        display_df["block_id"].astype(str)
    )
    for column in [
        "event_name",
        "cycle_goal",
        "scheduled_date",
        "start_date",
        "end_date",
        "training_category",
    ]:
        if column not in display_df.columns:
            display_df[column] = ""
    display_df["date_range"] = display_df.apply(date_range_label, axis=1)
    display_df = display_df.sort_values(
        ["week_num", "day_num", "start_date", "training_category", "id"],
        na_position="last",
    )
    if show_summary:
        visible_columns = [
            "date_range",
            "event_name",
            "cycle_goal",
            "week_num",
            "day_num",
            "training_category",
            "block",
            "notes",
            "created_at",
        ]
        visible_columns = [column for column in visible_columns if column in display_df.columns]
        st.dataframe(display_df[visible_columns], use_container_width=True, hide_index=True)

    if not show_period_view:
        return

    st.markdown("#### 週期檢視")
    for (week_num, day_num), day_df in display_df.groupby(["week_num", "day_num"], sort=True):
        first_row = day_df.iloc[0].to_dict()
        title_parts = [f"Week {week_num} / Day {day_num}"]
        date_text = date_range_label(first_row)
        if date_text:
            title_parts.append(date_text)
        if has_value(first_row.get("event_name")):
            title_parts.append(str(first_row.get("event_name")))

        with st.expander("｜".join(title_parts), expanded=False):
            if has_value(first_row.get("cycle_goal")):
                st.info(f"週期目標：{first_row.get('cycle_goal')}")
            for category, category_df in day_df.groupby("training_category", sort=False):
                st.markdown(f"**{category or '未分類'}**")
                for row in category_df.to_dict("records"):
                    notes = row.get("notes") or ""
                    block = blocks_by_id.get(row.get("block_id"))

                    if not show_block_details:
                        note_text = f"｜{notes}" if notes else ""
                        st.write(f"- {row.get('block')}{note_text}")
                        continue

                    with st.container(border=True):
                        st.markdown(f"##### {row.get('block')}")
                        meta_items = []
                        if has_value(row.get("event_name")):
                            meta_items.append(f"賽事/事件：{row.get('event_name')}")
                        row_date_text = date_range_label(row)
                        if row_date_text:
                            meta_items.append(f"日期：{row_date_text}")
                        if meta_items:
                            st.caption("｜".join(meta_items))
                        if notes:
                            st.info(f"教練備註：{notes}")

                        if not block:
                            st.warning("找不到這個板塊的詳細資料，可能已被刪除。")
                            continue

                        render_labeled_value("目標", block.get("goal"))
                        render_labeled_value("訓練元素", block.get("training_element"))
                        if block.get("description"):
                            st.caption(f"描述：{block.get('description')}")

                        render_assignment_detail_tables(
                            row["id"],
                            block["id"],
                            empty_message="這個板塊目前沒有建立詳細動作內容。",
                            show_report=show_report,
                            report_key_prefix="schedule_report",
                        )


def render_assignment_edit_form(
    row,
    block,
    form_key_prefix="edit_athlete_block_content",
):
    assignment_id = to_plain_int(row.get("id"))
    current_block_id = to_plain_int(row.get("block_id"))
    block_name = block_label(block) if block else row.get("block") or f"Block {current_block_id}"

    if not block:
        st.warning("找不到這個板塊模板，可能已經被刪除。你仍然可以刪除這筆學員課表安排。")

    exercises_df, exercises_error = assignment_exercises_for_editing(
        assignment_id,
        current_block_id,
    )
    if exercises_error:
        render_block_assignment_setup_help(exercises_error)
    elif exercises_df.empty:
        st.info("這個板塊目前沒有詳細動作內容。")

    with st.form(f"{form_key_prefix}_{assignment_id}"):
        st.markdown("##### 安排設定")
        event_name = render_event_input(
            f"{form_key_prefix}_assignment_{assignment_id}",
            row.get("event_name"),
        )
        cycle_goal = st.text_area(
            "週期目標",
            value=row.get("cycle_goal") if has_value(row.get("cycle_goal")) else "",
            height=80,
            key=f"{form_key_prefix}_cycle_goal_{assignment_id}",
        )
        col1, col2, col3 = st.columns([1, 1, 2])
        week_num = col1.number_input(
            "Week",
            min_value=1,
            value=to_plain_int(row.get("week_num")) or 1,
            key=f"{form_key_prefix}_week_{assignment_id}",
        )
        day_num = col2.number_input(
            "Day",
            min_value=1,
            value=to_plain_int(row.get("day_num")) or 1,
            key=f"{form_key_prefix}_day_{assignment_id}",
        )
        start_value = row.get("start_date") if has_value(row.get("start_date")) else row.get("scheduled_date")
        end_value = row.get("end_date") if has_value(row.get("end_date")) else start_value
        date_range = col3.date_input(
            "日期區間",
            value=(parse_date_value(start_value), parse_date_value(end_value)),
            key=f"{form_key_prefix}_date_{assignment_id}",
        )
        current_category = row.get("training_category") or TRAINING_CATEGORIES[0]
        category_index = (
            TRAINING_CATEGORIES.index(current_category)
            if current_category in TRAINING_CATEGORIES
            else 0
        )
        training_category = col3.selectbox(
            "訓練分類",
            TRAINING_CATEGORIES,
            index=category_index,
            key=f"{form_key_prefix}_category_{assignment_id}",
        )
        notes = st.text_area(
            "教練備註",
            value=row.get("notes") or "",
            key=f"{form_key_prefix}_notes_{assignment_id}",
        )

        st.markdown("##### 板塊內容")
        section_tables = {}
        if not exercises_df.empty:
            sorted_exercises_df = exercises_df.sort_values(["section_order", "order_num"])
            section_names = sorted_exercises_df["section_name"].dropna().unique()
            for section_name in section_names:
                safe_section_name = section_name or "未命名區段"
                st.markdown(f"**{safe_section_name}**")
                section_df = sorted_exercises_df[
                    sorted_exercises_df["section_name"] == section_name
                ].copy()
                for column in ATHLETE_BLOCK_EXERCISE_COLUMNS:
                    if column not in section_df.columns:
                        section_df[column] = ""
                section_tables[safe_section_name] = st.data_editor(
                    section_df[ATHLETE_BLOCK_EXERCISE_COLUMNS],
                    key=f"{form_key_prefix}_exercises_{assignment_id}_{safe_section_name}",
                    num_rows="dynamic",
                    use_container_width=True,
                    hide_index=True,
                    column_order=ATHLETE_BLOCK_EXERCISE_COLUMNS,
                    column_config=exercise_editor_column_config(),
                )

        submitted = st.form_submit_button("儲存這個學員的板塊內容", use_container_width=True)

    if submitted:
        try:
            start_date, end_date = normalize_date_range(date_range)
            update_athlete_block_assignment(
                assignment_id,
                current_block_id,
                event_name,
                cycle_goal,
                start_date,
                end_date,
                week_num,
                day_num,
                training_category,
                notes,
            )
            save_assignment_exercise_tables(assignment_id, section_tables)
            st.success("已更新這位學員的板塊內容。")
            st.rerun()
        except Exception as exc:
            render_block_assignment_setup_help(exc)

    if st.button("刪除這筆安排", key=f"{form_key_prefix}_ask_delete_{assignment_id}", use_container_width=True):
        st.session_state["pending_delete_assignment_id"] = assignment_id

    if st.session_state.get("pending_delete_assignment_id") == assignment_id:
        render_delete_athlete_block_assignment_confirmation(assignment_id, block_name)


def render_delete_athlete_block_assignment_confirmation(assignment_id, block_name):
    with st.container(border=True):
        st.warning(f"確認要從這位學員課表移除 {block_name} 嗎？板塊模板本身不會被刪除。")
        col1, col2 = st.columns(2)
        if col1.button("確認刪除", key=f"confirm_delete_assignment_{assignment_id}", use_container_width=True):
            try:
                delete_athlete_block_assignment(assignment_id)
                st.session_state.pop("pending_delete_assignment_id", None)
                st.success("已刪除這筆課表安排。")
                st.rerun()
            except Exception as exc:
                render_block_assignment_setup_help(exc)
        if col2.button("取消", key=f"cancel_delete_assignment_{assignment_id}", use_container_width=True):
            st.session_state.pop("pending_delete_assignment_id", None)
            st.rerun()


def find_logged_in_athlete(athletes_df):
    user_id = st.session_state.get("auth_user_id")
    email = st.session_state.get("auth_email")
    email_rows = pd.DataFrame()

    if "email" in athletes_df.columns and email:
        email_rows = athletes_df[
            athletes_df["email"].fillna("").str.lower() == email.lower()
        ]

    if "user_id" in athletes_df.columns and user_id:
        user_rows = athletes_df[
            athletes_df["user_id"].fillna("").astype(str) == str(user_id)
        ]
        if not user_rows.empty:
            matched_by_user = user_rows.iloc[0].to_dict()
            athlete_blocks_df, athlete_blocks_error = fetch_athlete_blocks(matched_by_user["id"])
            if athlete_blocks_error or not athlete_blocks_df.empty or email_rows.empty:
                return matched_by_user

    if not email_rows.empty:
        for _, athlete_row in email_rows.iterrows():
            athlete_blocks_df, athlete_blocks_error = fetch_athlete_blocks(athlete_row["id"])
            if not athlete_blocks_error and not athlete_blocks_df.empty:
                return athlete_row.to_dict()
        return email_rows.iloc[0].to_dict()

    return None


def student_page():
    restore_auth_session()
    if "auth_user_id" not in st.session_state:
        login_page()
        return

    st.title("我的課表")
    st.caption("查看自己目前被安排的訓練板塊")
    st.sidebar.caption(st.session_state.get("auth_email", ""))
    render_change_password_section()
    logout_button()

    athletes_df, athletes_error = fetch_athletes()
    if athletes_error:
        render_setup_help(athletes_error)
        return

    if athletes_df.empty:
        st.info("目前尚未建立學員。")
        return

    selected_athlete = find_logged_in_athlete(athletes_df)
    if not selected_athlete:
        st.warning("找不到和這個登入帳號對應的學員資料。")
        st.info("請確認教練新增學員時填寫的 Email 和你的登入 Email 相同。")
        return

    must_change_password = selected_athlete.get("must_change_password") is True
    if must_change_password:
        force_change_password_page(selected_athlete)
        return

    with st.container(border=True):
        st.markdown(f"### {selected_athlete.get('name') or '未命名學員'}")
        col1, col2, col3 = st.columns(3)
        col1.write(f"Email：{selected_athlete.get('email') or '-'}")
        col2.write(f"運動項目：{selected_athlete.get('sport') or '-'}")
        col3.write(f"學員 ID：{selected_athlete.get('id') or '-'}")

    blocks_df, blocks_error = fetch_blocks()
    if blocks_error:
        render_block_setup_help(blocks_error)
        return

    athlete_blocks_df, athlete_blocks_error = fetch_athlete_blocks(selected_athlete["id"])
    if athlete_blocks_error:
        render_block_assignment_setup_help(athlete_blocks_error)
        return

    if athlete_blocks_df.empty:
        st.info("目前還沒有被安排任何板塊。")
        st.caption(
            f"目前登入帳號對應到 athletes.id = {selected_athlete.get('id')}。"
            "請確認 Supabase 的 athlete_blocks.athlete_id 是否等於這個 ID。"
        )
    else:
        render_student_schedule_calendar(
            selected_athlete["id"],
            athlete_blocks_df,
            blocks_df,
        )


def update_current_user_password(new_password):
    supabase.auth.update_user({"password": new_password})


def force_change_password_page(selected_athlete):
    st.title("請設定新密碼")
    st.caption("你目前使用的是臨時密碼。設定新密碼後才能查看課表。")

    with st.form("force_change_password_form"):
        new_password = st.text_input("新 Password", type="password")
        confirm_password = st.text_input("確認新 Password", type="password")
        submitted = st.form_submit_button("更新密碼", use_container_width=True)

    if submitted:
        if not new_password:
            st.warning("請輸入新 Password。")
        elif len(new_password) < 6:
            st.warning("Password 至少需要 6 碼。")
        elif new_password != confirm_password:
            st.warning("兩次輸入的 Password 不一致。")
        else:
            try:
                update_current_user_password(new_password)
                supabase.table("athletes").update(
                    {"must_change_password": False}
                ).eq("id", selected_athlete["id"]).execute()
                st.success("密碼已更新。")
                st.rerun()
            except Exception as exc:
                st.error(f"更新密碼失敗：{exc}")


def render_change_password_section():
    with st.sidebar.expander("修改密碼", expanded=False):
        st.caption("需要時再打開修改。")
        with st.form("change_password_form"):
            new_password = st.text_input("新 Password", type="password")
            confirm_password = st.text_input("確認新 Password", type="password")
            submitted = st.form_submit_button("更新密碼", use_container_width=True)

        if submitted:
            if not new_password:
                st.warning("請輸入新 Password。")
            elif len(new_password) < 6:
                st.warning("Password 至少需要 6 碼。")
            elif new_password != confirm_password:
                st.warning("兩次輸入的 Password 不一致。")
            else:
                try:
                    update_current_user_password(new_password)
                    st.success("密碼已更新。下次請使用新密碼登入。")
                except Exception as exc:
                    st.error(f"更新密碼失敗：{exc}")


def main():
    inject_responsive_styles()

    if password_recovery_page():
        return

    mode = st.sidebar.radio("介面", ["教練端", "學員端"])

    if mode == "學員端":
        student_page()
        return

    page = st.sidebar.radio("教練功能", ["學員", "板塊"])
    if page == "學員":
        coach_page()
    else:
        blocks_page()


main()
